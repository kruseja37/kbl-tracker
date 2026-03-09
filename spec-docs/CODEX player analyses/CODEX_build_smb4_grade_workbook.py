#!/usr/bin/env python3
"""Build an XLSX workbook for SMB4 grade calculation and player generation.

Output workbook is designed for direct upload/import into Google Sheets.
The calculator mirrors the improved baked grade model used by the toolkit.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT_PATH = "/Users/johnkruse/Projects/kbl-tracker/spec-docs/CODEX_SMB4_Grade_Calculator_and_Generator.xlsx"

FULL_GRADE_SCALE = [
    "S", "A+", "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "E+", "E", "E-", "F"
]
GRADE_TO_POINTS = {g: len(FULL_GRADE_SCALE) - 1 - i for i, g in enumerate(FULL_GRADE_SCALE)}
GRADE_NUMERIC_CENTERS = {
    "S": 97.0,
    "A+": 92.0,
    "A": 87.0,
    "A-": 82.0,
    "B+": 77.0,
    "B": 72.0,
    "B-": 67.0,
    "C+": 62.0,
    "C": 57.0,
    "C-": 52.0,
    "D+": 47.0,
    "D": 42.0,
    "D-": 37.0,
    "E+": 32.0,
    "E": 27.0,
    "E-": 22.0,
    "F": 15.0,
}

GRADE_LOWER_BOUNDS = [
    (0.0, 0, "F", 15.0),
    (18.5, 1, "E-", 22.0),
    (24.5, 2, "E", 27.0),
    (29.5, 3, "E+", 32.0),
    (34.5, 4, "D-", 37.0),
    (39.5, 5, "D", 42.0),
    (44.5, 6, "D+", 47.0),
    (49.5, 7, "C-", 52.0),
    (54.5, 8, "C", 57.0),
    (59.5, 9, "C+", 62.0),
    (64.5, 10, "B-", 67.0),
    (69.5, 11, "B", 72.0),
    (74.5, 12, "B+", 77.0),
    (79.5, 13, "A-", 82.0),
    (84.5, 14, "A", 87.0),
    (89.5, 15, "A+", 92.0),
    (94.5, 16, "S", 97.0),
]

VERSATILITY_MAP = {
    "(none)": 0,
    "1B": 1,
    "1B/OF": 4,
    "2B": 1,
    "3B": 1,
    "C": 1,
    "IF": 4,
    "IF/OF": 7,
    "LF": 1,
    "OF": 3,
    "RF": 1,
    "SS": 1,
}

POSITIVE_TRAITS = {
    "Cannon Arm", "Durable", "First Pitch Slayer", "Sprinter", "K Collector", "Tough Out",
    "Stimulated", "Specialist", "Reverse Splits", "Stealer", "Pick Officer", "Sign Stealer",
    "Mind Gamer", "Distractor", "Bad Ball Hitter", "Pinch Perfect", "Base Rounder", "Composed",
    "Magic Hands", "Fastball Hitter", "Off-Speed Hitter", "Low Pitch", "High Pitch", "Inside Pitch",
    "Outside Pitch", "Metal Head", "Consistent", "Two Way", "Rally Stopper", "Clutch", "Dive Wizard",
    "Rally Starter", "RBI Hero", "CON vs LHP", "CON vs RHP", "POW vs LHP", "POW vs RHP",
    "Ace Exterminator", "Bunter", "Utility", "Big Hack", "Little Hack", "Gets Ahead",
    "Elite 4F", "Elite 2F", "Elite CF", "Elite FK", "Elite SL", "Elite CB", "Elite CH", "Elite SB",
}

NEGATIVE_TRAITS = {
    "K Neglecter", "K Neglector", "Whiffer", "Slow Poke", "First Pitch Prayer", "Injury Prone", "Noodle Arm",
    "Bad Jumps", "Easy Jumps", "Wild Thrower", "Easy Target", "Base Jogger", "BB Prone",
    "Butter Fingers", "Volatile", "Choker", "Meltdown", "Surrounded", "Wild Thing", "RBI Zero",
    "Falls Behind", "Crossed Up",
}

HITTER_PRIMARY_COEF = {
    "1B": 0.0,
    "2B": 0.8313525554,
    "3B": -1.2668027922,
    "C": 2.2997744611,
    "CF": 0.6637032249,
    "LF": -0.5614229268,
    "RF": -0.1963177907,
    "SS": 0.1115302985,
}

HITTER_SECONDARY_COEF = {
    "(none)": 0.0,
    "1B": -0.0743925098,
    "1B/OF": -0.9164565494,
    "2B": 0.8398362231,
    "3B": 0.3035512464,
    "C": -0.8721695187,
    "IF": 0.5288517627,
    "IF/OF": 0.0,
    "LF": 0.0694935796,
    "OF": 0.1776157661,
    "RF": -0.4608455812,
    "SS": 0.7115980783,
}

PITCHER_PRIMARY_COEF = {
    "CP": 0.0,
    "RP": 0.0143423869,
    "SP": 0.2361613210,
    "SP/RP": -1.0149995619,
}

HITTER_TRAIT_COEF = {
    "First Pitch Slayer": 0.6985887989,
    "Little Hack": -1.2708309640,
    "Mind Gamer": 1.5014276130,
    "Rally Starter": -0.1998738040,
    "Magic Hands": -0.8310486121,
    "Utility": -0.3052216719,
    "Big Hack": -0.2192440495,
    "Sprinter": -0.2816825373,
    "Cannon Arm": -0.5417638949,
    "Fastball Hitter": 2.4542565444,
    "Bad Ball Hitter": -0.4592940132,
    "Whiffer": -0.7381706008,
}

PITCHER_TRAIT_COEF = {
    "K Collector": 0.9087461920,
    "Gets Ahead": 1.0320910791,
    "Elite 2F": -0.5280484145,
    "Elite 4F": 0.4167563425,
    "Falls Behind": -0.9976052263,
    "Elite CF": 0.7579419877,
    "Rally Stopper": -1.1328476477,
    "Elite FK": 0.4854804797,
    "Specialist": 2.1540724826,
    "Crossed Up": 1.5628438611,
    "Elite CB": -0.2319264114,
    "Volatile": 1.6516084637,
}

PITCH_COEF = {
    "2F": 0.4731924690,
    "4F": 0.6321855453,
    "CB": 0.1994458286,
    "CF": 0.4593316468,
    "CH": 0.0411500855,
    "FK": -0.4882808759,
    "SB": 0.1040401887,
    "SL": -0.4119626453,
}

ARCHETYPE_BLUEPRINTS = {
    "balanced": {"C": 2, "1B": 2, "2B": 2, "3B": 1, "SS": 1, "LF": 2, "CF": 1, "RF": 2, "SP": 4, "SP/RP": 1, "RP": 3, "CP": 1},
    "power-heavy": {"C": 2, "1B": 3, "2B": 1, "3B": 2, "SS": 1, "LF": 2, "CF": 1, "RF": 2, "SP": 4, "SP/RP": 1, "RP": 2, "CP": 1},
    "speed-defense": {"C": 2, "1B": 1, "2B": 2, "3B": 1, "SS": 2, "LF": 1, "CF": 2, "RF": 1, "SP": 4, "SP/RP": 1, "RP": 4, "CP": 1},
    "bullpen-heavy": {"C": 2, "1B": 1, "2B": 2, "3B": 1, "SS": 2, "LF": 1, "CF": 1, "RF": 1, "SP": 3, "SP/RP": 2, "RP": 5, "CP": 1},
}


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, min_col=1, max_col=26, width=16):
    for c in range(min_col, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = width


def write_kv_sheet(wb, title, headers, items, widths=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in items:
        ws.append(list(row))
    style_header(ws)
    if widths:
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    else:
        autosize(ws, 1, len(headers), 18)
    return ws


def make_readme(ws):
    ws.title = "README"
    lines = [
        "SMB4 Grade Calculator + Fictional Player Generator",
        "",
        "This workbook is designed for upload to Google Sheets.",
        "",
        "Tabs:",
        "- Calculator: exact improved scorer for manual player inputs.",
        "- Generator: heuristic random generator that scores the result with the same exact model.",
        "- Archetype_Presets: roster templates for balanced/power/speed-defense/bullpen-heavy teams.",
        "- Lookup tabs: grade thresholds, trait polarity, versatility, position coefficients, trait coefficients, pitch coefficients.",
        "",
        "Model factors included:",
        "- Primary position",
        "- Secondary position and combo positions",
        "- Bats / Throws",
        "- Positive and negative traits",
        "- Full ratings",
        "- Pitcher batting stats",
        "- Pitch arsenal count and pitch-type coefficients for pitchers",
        "",
        "Important note:",
        "- The Calculator tab matches the improved baked model.",
        "- The Generator tab is a heuristic starter; final grade shown there is exact for the generated line.",
    ]
    for i, line in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=line)
    ws["A1"].font = Font(bold=True, size=14)
    ws.column_dimensions["A"].width = 118


def make_grade_scale(wb):
    ws = wb.create_sheet("Grade_Scale")
    ws.append(["LowerBound", "GradeIdx", "Grade", "Center", "GradeIdxByGrade"])
    for lower_bound, idx, grade, center in GRADE_LOWER_BOUNDS:
        ws.append([lower_bound, idx, grade, center, idx])
    style_header(ws)
    autosize(ws, 1, 5, 14)


def make_trait_polarity(wb):
    ws = wb.create_sheet("Trait_Polarity")
    ws.append(["Trait", "PosFlag", "NegFlag", "Polarity"])
    all_traits = sorted(POSITIVE_TRAITS.union(NEGATIVE_TRAITS))
    for trait in all_traits:
        pos = 1 if trait in POSITIVE_TRAITS else 0
        neg = 1 if trait in NEGATIVE_TRAITS else 0
        ws.append([trait, pos, neg, "Positive" if pos else "Negative"])
    style_header(ws)
    autosize(ws, 1, 4, 20)


def make_lookup_tables(wb):
    write_kv_sheet(
        wb,
        "Secondary_Versatility",
        ["SecondaryPos", "Versatility"],
        sorted(VERSATILITY_MAP.items()),
        widths={"A": 16, "B": 12},
    )
    write_kv_sheet(
        wb,
        "Hitter_Primary_Coef",
        ["PrimaryPos", "Coef"],
        sorted(HITTER_PRIMARY_COEF.items()),
        widths={"A": 16, "B": 12},
    )
    write_kv_sheet(
        wb,
        "Hitter_Secondary_Coef",
        ["SecondaryPos", "Coef"],
        sorted(HITTER_SECONDARY_COEF.items()),
        widths={"A": 16, "B": 12},
    )
    write_kv_sheet(
        wb,
        "Pitcher_Primary_Coef",
        ["PrimaryPos", "Coef"],
        sorted(PITCHER_PRIMARY_COEF.items()),
        widths={"A": 16, "B": 12},
    )
    write_kv_sheet(
        wb,
        "Hitter_Trait_Coef",
        ["Trait", "Coef"],
        sorted(HITTER_TRAIT_COEF.items()),
        widths={"A": 22, "B": 12},
    )
    write_kv_sheet(
        wb,
        "Pitcher_Trait_Coef",
        ["Trait", "Coef"],
        sorted(PITCHER_TRAIT_COEF.items()),
        widths={"A": 22, "B": 12},
    )
    write_kv_sheet(
        wb,
        "Pitch_Coef",
        ["Pitch", "Coef"],
        sorted(PITCH_COEF.items()),
        widths={"A": 10, "B": 12},
    )


def make_archetype_presets(wb):
    ws = wb.create_sheet("Archetype_Presets")
    ws.append([
        "Archetype", "C", "1B", "2B", "3B", "SS", "LF", "CF", "RF", "SP", "SP/RP", "RP", "CP", "Total"
    ])
    for archetype, counts in ARCHETYPE_BLUEPRINTS.items():
        ws.append([
            archetype,
            counts["C"], counts["1B"], counts["2B"], counts["3B"], counts["SS"],
            counts["LF"], counts["CF"], counts["RF"], counts["SP"], counts["SP/RP"], counts["RP"], counts["CP"],
            sum(counts.values()),
        ])
    style_header(ws)
    autosize(ws, 1, 14, 14)
    ws.column_dimensions["A"].width = 18


def calculator_numeric_formula(r: int) -> str:
    hitter_formula = (
        f'10.5965166711+0.2825983581*I{r}+0.2806503532*J{r}+0.2027213083*K{r}+0.1147824982*L{r}+0.0915305332*M{r}'
        f'-0.0088454122*(I{r}*J{r}/100)-0.0336045706*(K{r}*L{r}/100)'
        f'+2.8497389733*IF(D{r}="L",1,0)+4.5116226727*IF(D{r}="S",1,0)-0.6571546448*IF(E{r}="L",1,0)'
        f'+0.0850728147*T{r}+0.0129488446*(T{r}*T{r})+0.1909373936*IF(OR(F{r}="Utility",G{r}="Utility"),T{r},0)'
        f'+0.9656824071*R{r}-1.7517683256*S{r}+W{r}+X{r}+Y{r}'
    )
    pitcher_formula = (
        f'16.5944849573+0.2529999141*N{r}+0.2665900378*O{r}+0.2632687105*P{r}'
        f'+0.0427586837*I{r}+0.0534777057*J{r}+0.009015832*K{r}+0.0204898106*(O{r}*P{r}/100)'
        f'+1.0091022427*U{r}+1.0968542297*IF(D{r}="L",1,0)+0.3771555045*IF(D{r}="S",1,0)-0.2226159177*IF(E{r}="L",1,0)'
        f'+1.21385024*R{r}-1.1652812274*S{r}+W{r}+X{r}+Z{r}'
    )
    return f'=IF(Q{r}="P",{pitcher_formula},{hitter_formula})'


def make_calculator(wb):
    ws = wb.create_sheet("Calculator")
    headers = [
        "Name", "PrimaryPos", "SecondaryPos", "Bats", "Throws", "Trait1", "Trait2", "Arsenal",
        "POW", "CON", "SPD", "FLD", "ARM", "VEL", "JNK", "ACC",
        "PlayerType", "PosTraits", "NegTraits", "Versatility", "ArsenalCount", "BaseWeighted",
        "TraitCoef", "PrimaryCoef", "SecondaryCoef", "PitchCoef", "NumericScore", "GradeIdx", "Grade",
        "TargetGrade", "DeltaIdx (Pred-Target)", "Notes",
    ]
    ws.append(headers)
    style_header(ws)
    ws.freeze_panes = "A2"

    for r in range(2, 1002):
        ws[f"Q{r}"] = f'=IF(OR(B{r}="SP",B{r}="RP",B{r}="CP",B{r}="SP/RP"),"P","H")'
        ws[f"R{r}"] = f'=IF(F{r}="",0,IFNA(VLOOKUP(F{r},Trait_Polarity!$A:$C,2,FALSE),0))+IF(G{r}="",0,IFNA(VLOOKUP(G{r},Trait_Polarity!$A:$C,2,FALSE),0))'
        ws[f"S{r}"] = f'=IF(F{r}="",0,IFNA(VLOOKUP(F{r},Trait_Polarity!$A:$C,3,FALSE),0))+IF(G{r}="",0,IFNA(VLOOKUP(G{r},Trait_Polarity!$A:$C,3,FALSE),0))'
        ws[f"T{r}"] = f'=IF(Q{r}="P",0,IFNA(VLOOKUP(IF(C{r}="","(none)",C{r}),Secondary_Versatility!$A:$B,2,FALSE),0))'
        ws[f"U{r}"] = f'=IF(H{r}="",0,LEN(H{r})-LEN(SUBSTITUTE(H{r},"|",""))+1)'
        ws[f"V{r}"] = f'=IF(Q{r}="P",(N{r}+O{r}+P{r})/3,0.3*I{r}+0.3*J{r}+0.2*K{r}+0.1*L{r}+0.1*M{r})'
        ws[f"W{r}"] = (
            f'=IF(Q{r}="P",'
            f'IF(F{r}="",0,IFNA(VLOOKUP(F{r},Pitcher_Trait_Coef!$A:$B,2,FALSE),0))+IF(G{r}="",0,IFNA(VLOOKUP(G{r},Pitcher_Trait_Coef!$A:$B,2,FALSE),0)),'
            f'IF(F{r}="",0,IFNA(VLOOKUP(F{r},Hitter_Trait_Coef!$A:$B,2,FALSE),0))+IF(G{r}="",0,IFNA(VLOOKUP(G{r},Hitter_Trait_Coef!$A:$B,2,FALSE),0)))'
        )
        ws[f"X{r}"] = f'=IF(Q{r}="P",IFNA(VLOOKUP(B{r},Pitcher_Primary_Coef!$A:$B,2,FALSE),0),IFNA(VLOOKUP(B{r},Hitter_Primary_Coef!$A:$B,2,FALSE),0))'
        ws[f"Y{r}"] = f'=IF(Q{r}="P",0,IFNA(VLOOKUP(IF(C{r}="","(none)",C{r}),Hitter_Secondary_Coef!$A:$B,2,FALSE),0))'
        ws[f"Z{r}"] = (
            f'=IF(Q{r}="P",'
            f'IF(ISNUMBER(SEARCH("2F",H{r})),IFNA(VLOOKUP("2F",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
            f'IF(ISNUMBER(SEARCH("4F",H{r})),IFNA(VLOOKUP("4F",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
            f'IF(ISNUMBER(SEARCH("CB",H{r})),IFNA(VLOOKUP("CB",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
            f'IF(ISNUMBER(SEARCH("CF",H{r})),IFNA(VLOOKUP("CF",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
            f'IF(ISNUMBER(SEARCH("CH",H{r})),IFNA(VLOOKUP("CH",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
            f'IF(ISNUMBER(SEARCH("FK",H{r})),IFNA(VLOOKUP("FK",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
            f'IF(ISNUMBER(SEARCH("SB",H{r})),IFNA(VLOOKUP("SB",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
            f'IF(ISNUMBER(SEARCH("SL",H{r})),IFNA(VLOOKUP("SL",Pitch_Coef!$A:$B,2,FALSE),0),0),0)'
        )
        ws[f"AA{r}"] = calculator_numeric_formula(r)
        ws[f"AB{r}"] = f'=LOOKUP(AA{r},Grade_Scale!$A$2:$A$18,Grade_Scale!$B$2:$B$18)'
        ws[f"AC{r}"] = f'=LOOKUP(AA{r},Grade_Scale!$A$2:$A$18,Grade_Scale!$C$2:$C$18)'
        ws[f"AE{r}"] = f'=IF(AD{r}="","",AB{r}-VLOOKUP(AD{r},Grade_Scale!$C$2:$E$18,3,FALSE))'

    widths = {
        "A": 18, "B": 12, "C": 14, "D": 8, "E": 8, "F": 18, "G": 18, "H": 20,
        "I": 7, "J": 7, "K": 7, "L": 7, "M": 7, "N": 7, "O": 7, "P": 7,
        "Q": 10, "R": 10, "S": 10, "T": 11, "U": 12, "V": 12, "W": 10, "X": 10,
        "Y": 12, "Z": 10, "AA": 12, "AB": 10, "AC": 8, "AD": 10, "AE": 18, "AF": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def make_generator(wb):
    ws = wb.create_sheet("Generator")
    ws["A1"] = "Fictional Player Generator"
    ws["A1"].font = Font(bold=True, size=13)

    labels = [
        (3, "Mode (H or P)"),
        (4, "TargetGrade"),
        (5, "PrimaryPos"),
        (6, "SecondaryPos"),
        (7, "Bats"),
        (8, "Throws"),
        (9, "Trait1"),
        (10, "Trait2"),
        (11, "Arsenal (optional)"),
        (12, "Pitcher BatPOW"),
        (13, "Pitcher BatCON"),
        (14, "Pitcher BatSPD"),
        (16, "TargetCenter"),
        (17, "PosTraits"),
        (18, "NegTraits"),
        (19, "Versatility"),
        (20, "TraitCoef"),
        (21, "PrimaryCoef"),
        (22, "SecondaryCoef"),
        (23, "PitchCoef"),
        (24, "ArsenalCount"),
        (25, "EstimatedBase"),
    ]
    for row, label in labels:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(bold=True)

    defaults = {
        "B3": "H",
        "B4": "B",
        "B5": "SS",
        "B6": "2B",
        "B7": "R",
        "B8": "R",
        "B9": "",
        "B10": "",
        "B11": "",
        "B12": 15,
        "B13": 20,
        "B14": 35,
    }
    for cell, value in defaults.items():
        ws[cell] = value

    ws["B16"] = '=VLOOKUP(B4,Grade_Scale!$C$2:$D$18,2,FALSE)'
    ws["B17"] = '=IF(B9="",0,IFNA(VLOOKUP(B9,Trait_Polarity!$A:$C,2,FALSE),0))+IF(B10="",0,IFNA(VLOOKUP(B10,Trait_Polarity!$A:$C,2,FALSE),0))'
    ws["B18"] = '=IF(B9="",0,IFNA(VLOOKUP(B9,Trait_Polarity!$A:$C,3,FALSE),0))+IF(B10="",0,IFNA(VLOOKUP(B10,Trait_Polarity!$A:$C,3,FALSE),0))'
    ws["B19"] = '=IF(B3="P",0,IFNA(VLOOKUP(IF(B6="","(none)",B6),Secondary_Versatility!$A:$B,2,FALSE),0))'
    ws["B20"] = (
        '=IF(B3="P",'
        'IF(B9="",0,IFNA(VLOOKUP(B9,Pitcher_Trait_Coef!$A:$B,2,FALSE),0))+IF(B10="",0,IFNA(VLOOKUP(B10,Pitcher_Trait_Coef!$A:$B,2,FALSE),0)),'
        'IF(B9="",0,IFNA(VLOOKUP(B9,Hitter_Trait_Coef!$A:$B,2,FALSE),0))+IF(B10="",0,IFNA(VLOOKUP(B10,Hitter_Trait_Coef!$A:$B,2,FALSE),0)))'
    )
    ws["B21"] = '=IF(B3="P",IFNA(VLOOKUP(B5,Pitcher_Primary_Coef!$A:$B,2,FALSE),0),IFNA(VLOOKUP(B5,Hitter_Primary_Coef!$A:$B,2,FALSE),0))'
    ws["B22"] = '=IF(B3="P",0,IFNA(VLOOKUP(IF(B6="","(none)",B6),Hitter_Secondary_Coef!$A:$B,2,FALSE),0))'
    ws["B23"] = (
        '=IF(B3="P",'
        'IF(ISNUMBER(SEARCH("2F",B11)),IFNA(VLOOKUP("2F",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
        'IF(ISNUMBER(SEARCH("4F",B11)),IFNA(VLOOKUP("4F",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
        'IF(ISNUMBER(SEARCH("CB",B11)),IFNA(VLOOKUP("CB",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
        'IF(ISNUMBER(SEARCH("CF",B11)),IFNA(VLOOKUP("CF",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
        'IF(ISNUMBER(SEARCH("CH",B11)),IFNA(VLOOKUP("CH",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
        'IF(ISNUMBER(SEARCH("FK",B11)),IFNA(VLOOKUP("FK",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
        'IF(ISNUMBER(SEARCH("SB",B11)),IFNA(VLOOKUP("SB",Pitch_Coef!$A:$B,2,FALSE),0),0)+'
        'IF(ISNUMBER(SEARCH("SL",B11)),IFNA(VLOOKUP("SL",Pitch_Coef!$A:$B,2,FALSE),0),0),0)'
    )
    ws["B24"] = '=IF(B11="",0,LEN(B11)-LEN(SUBSTITUTE(B11,"|",""))+1)'
    ws["B25"] = '=IF(B3="P",MAX(18,MIN(96,B16-15)),MAX(15,MIN(92,B16-10)))'

    ws["A28"] = "Generated Ratings"
    ws["A28"].font = Font(bold=True)
    gen_labels = ["POW", "CON", "SPD", "FLD", "ARM", "VEL", "JNK", "ACC", "NumericScore", "GradeIdx", "Grade"]
    for offset, label in enumerate(gen_labels, start=29):
        ws[f"A{offset}"] = label
        ws[f"A{offset}"].font = Font(bold=True)

    ws["C29"] = '=IF($B$5="1B",15,IF($B$5="2B",-10,IF($B$5="SS",-10,IF($B$5="3B",10,IF($B$5="LF",10,IF($B$5="CF",-10,IF($B$5="RF",5,0)))))))'
    ws["C30"] = '=IF($B$5="2B",5,0)'
    ws["C31"] = '=IF($B$5="C",-10,IF($B$5="1B",-10,IF($B$5="2B",5,IF($B$5="SS",5,IF($B$5="3B",-10,IF($B$5="CF",15,IF($B$5="RF",-5,0)))))))'
    ws["C32"] = '=IF($B$5="C",10,IF($B$5="1B",-5,IF($B$5="SS",10,IF($B$5="LF",-5,IF($B$5="CF",5,0)))))'
    ws["C33"] = '=IF($B$5="C",10,IF($B$5="SS",5,IF($B$5="3B",5,IF($B$5="LF",-5,IF($B$5="RF",10,0)))))'
    ws["C34"] = '=IF($B$5="SP",-2,IF($B$5="CP",8,IF($B$5="SP/RP",3,2)))'
    ws["C35"] = '=IF($B$5="SP",-3,IF($B$5="CP",5,IF($B$5="SP/RP",2,1)))'
    ws["C36"] = '=IF($B$5="SP",5,IF($B$5="CP",-13,IF($B$5="SP/RP",0,-2)))'

    ws["B29"] = '=IF($B$3="H",MAX(0,MIN(99,ROUND($B$25+C29+(RAND()-0.5)*18,0))),$B$12)'
    ws["B30"] = '=IF($B$3="H",MAX(0,MIN(99,ROUND($B$25+C30+(RAND()-0.5)*18,0))),$B$13)'
    ws["B31"] = '=IF($B$3="H",MAX(0,MIN(99,ROUND($B$25+C31+(RAND()-0.5)*18,0))),$B$14)'
    ws["B32"] = '=IF($B$3="H",MAX(0,MIN(99,ROUND($B$25+C32+(RAND()-0.5)*18,0))),ROUND(40+RAND()*30,0))'
    ws["B33"] = '=IF($B$3="H",MAX(0,MIN(99,ROUND($B$25+C33+(RAND()-0.5)*18,0))),0)'
    ws["B34"] = '=IF($B$3="P",MAX(0,MIN(99,ROUND($B$25+C34+(RAND()-0.5)*18,0))),0)'
    ws["B35"] = '=IF($B$3="P",MAX(0,MIN(99,ROUND($B$25+C35+(RAND()-0.5)*18,0))),0)'
    ws["B36"] = '=IF($B$3="P",MAX(0,MIN(99,ROUND($B$25+C36+(RAND()-0.5)*18,0))),0)'
    ws["B37"] = (
        '=IF($B$3="P",'
        '16.5944849573+0.2529999141*B34+0.2665900378*B35+0.2632687105*B36+0.0427586837*B29+0.0534777057*B30+0.009015832*B31+0.0204898106*(B35*B36/100)'
        '+1.0091022427*$B$24+1.0968542297*IF($B$7="L",1,0)+0.3771555045*IF($B$7="S",1,0)-0.2226159177*IF($B$8="L",1,0)'
        '+1.21385024*$B$17-1.1652812274*$B$18+$B$20+$B$21+$B$23,'
        '10.5965166711+0.2825983581*B29+0.2806503532*B30+0.2027213083*B31+0.1147824982*B32+0.0915305332*B33'
        '-0.0088454122*(B29*B30/100)-0.0336045706*(B31*B32/100)+2.8497389733*IF($B$7="L",1,0)+4.5116226727*IF($B$7="S",1,0)-0.6571546448*IF($B$8="L",1,0)'
        '+0.0850728147*$B$19+0.0129488446*($B$19*$B$19)+0.1909373936*IF(OR($B$9="Utility",$B$10="Utility"),$B$19,0)'
        '+0.9656824071*$B$17-1.7517683256*$B$18+$B$20+$B$21+$B$22)'
    )
    ws["B38"] = '=LOOKUP(B37,Grade_Scale!$A$2:$A$18,Grade_Scale!$B$2:$B$18)'
    ws["B39"] = '=LOOKUP(B37,Grade_Scale!$A$2:$A$18,Grade_Scale!$C$2:$C$18)'

    ws["D3"] = "Use the Generator tab to create a random starting player. The grade shown here is exact for the generated ratings."
    ws["D4"] = "For pitchers, type an arsenal like 4F|CF|SL|CH to include pitch-count and pitch-type effects."
    ws["D5"] = "If you need exact custom scoring, copy the generated line into Calculator and edit manually."

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 96


def main():
    wb = Workbook()
    make_readme(wb.active)
    make_grade_scale(wb)
    make_trait_polarity(wb)
    make_lookup_tables(wb)
    make_archetype_presets(wb)
    make_calculator(wb)
    make_generator(wb)
    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
