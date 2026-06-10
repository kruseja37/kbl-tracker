#!/usr/bin/env python3
"""
extract-iv-data.py — T1: IV Curve & Trait Pricing Data Extraction

Reads the XBL workbook (spec-docs/reference/Team_Builder_Archetype_Logic_Template.xlsx)
and deterministically regenerates two typed TypeScript data files:

  src/data/ivCurves.ts      — Salary Cap sheet curve params (18 position blocks,
                              primary cols C–H + sub-minimum reverse cols I–N)
  src/data/traitPricing.ts  — Traits sheet (75 traits: deltas, multipliers, flat fees)
                              + pitch costs, bullpen arsenal tax, aux pricing

Spec: spec-docs/IV_ENGINE_AND_ROSTER_INTELLIGENCE_SPEC.md §3.2–§3.6, §13 (T1)

Usage:  python3 scripts/extract-iv-data.py
Requires: openpyxl (extraction only — no runtime npm dependency).

The script HARD-FAILS (assert) if the workbook layout deviates from the spec's
described structure, rather than emitting silently wrong data.
"""

import datetime
import os
import re
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "spec-docs", "reference", "Team_Builder_Archetype_Logic_Template.xlsx")
OUT_CURVES = os.path.join(ROOT, "src", "data", "ivCurves.ts")
OUT_TRAITS = os.path.join(ROOT, "src", "data", "traitPricing.ts")

WORKBOOK_REL = "spec-docs/reference/Team_Builder_Archetype_Logic_Template.xlsx"
SPEC_REL = "spec-docs/IV_ENGINE_AND_ROSTER_INTELLIGENCE_SPEC.md"
TODAY = datetime.date.today().isoformat()

# Spec §3.3 — Lists!AN2:AO19 position → first attribute row of its "Salary Cap" block.
EXPECTED_POSITION_ROWS = [
    ("C", 5), ("1B", 11), ("2B", 17), ("SS", 23), ("3B", 29), ("LF", 35),
    ("CF", 41), ("RF", 47), ("IF", 53), ("OF", 59), ("IF/OF", 65), ("-", 71),
    ("SP", 77), ("SP/RP", 85), ("RP", 93), ("CP", 101), ("1B/OF", 109), ("EXTRA", 117),
]
HITTER_ATTRS = ["POW", "CON", "SPD", "FLD", "ARM"]
PITCHER_ATTRS = ["POW", "CON", "SPD", "FLD", "VEL", "JNK", "ACC"]
ALL_ATTRS = ["POW", "CON", "SPD", "FLD", "ARM", "VEL", "JNK", "ACC"]

# Spec §3.3 verification anchor: C-position primary curve params.
C_ANCHORS = {
    "POW": (0, 1, 50, 8000, 1.5, 56000),
    "CON": (0, 1, 55, 7000, 2, 31500),
    "SPD": (0, 1, 55, 5500, 3, 34000),
    "FLD": (0, 1, 60, 1400, 2, 5600),
    "ARM": (0, 1, 60, 2550, 2, 10200),
}

# §3.6 flagged blank cells ("·") to verify: (trait label, delta attr)
FLAGGED_BLANKS = [
    ("Bad Jumps (-)", "POW"), ("Big Hack (+)", "CON"), ("Little Hack (+)", "POW"),
    ("Crossed Up (-)", "POW"),
    ("Rally Stopper (+)", "POW"), ("Rally Stopper (+)", "CON"), ("Rally Stopper (+)", "SPD"),
    ("Rally Stopper (+)", "FLD"), ("Rally Stopper (+)", "ARM"),
    ("Easy Target (-)", "VEL"), ("Easy Target (-)", "JNK"), ("Easy Target (-)", "ACC"),
    ("Two Way (C) (+)", "FLD"), ("Two Way (OF) (+)", "ARM"),
    ("Wild Thrower (-)", "ACC"),
    ("Metal Head (+)", "POW"), ("Metal Head (+)", "CON"),
    ("Elite 2F (+)", "VEL"), ("Elite 2F (+)", "JNK"), ("Elite 2F (+)", "ACC"),
    ("Elite 4F (+)", "VEL"), ("Elite 4F (+)", "JNK"), ("Elite 4F (+)", "ACC"),
    ("Elite CB (+)", "VEL"), ("Elite CB (+)", "JNK"), ("Elite CB (+)", "ACC"),
    ("Elite CF (+)", "VEL"), ("Elite CF (+)", "JNK"), ("Elite CF (+)", "ACC"),
    ("Elite CH (+)", "VEL"), ("Elite CH (+)", "JNK"), ("Elite CH (+)", "ACC"),
    ("Elite FK (+)", "VEL"), ("Elite FK (+)", "JNK"), ("Elite FK (+)", "ACC"),
    ("Elite SB (+)", "VEL"), ("Elite SB (+)", "JNK"), ("Elite SB (+)", "ACC"),
    ("Elite SL (+)", "VEL"), ("Elite SL (+)", "JNK"), ("Elite SL (+)", "ACC"),
    ("Reverse Splits (+)", "VEL"), ("Reverse Splits (+)", "JNK"), ("Reverse Splits (+)", "ACC"),
    ("Specialist (+)", "VEL"), ("Specialist (+)", "JNK"), ("Specialist (+)", "ACC"),
]


def num(v):
    """Format a cell number for TS output: drop trailing .0, keep real decimals."""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return repr(v)


def is_number(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def fail(msg):
    print(f"EXTRACTION FAILED: {msg}", file=sys.stderr)
    sys.exit(1)


def extract_curves(wb):
    """Salary Cap sheet → 18 position blocks of {attr: {primary, subMin?}}."""
    lists = wb["Lists"]
    actual_map = [(lists.cell(row=r, column=40).value, lists.cell(row=r, column=41).value)
                  for r in range(2, 20)]  # AN=40, AO=41
    actual_map = [(k, int(v)) for k, v in actual_map]
    if actual_map != EXPECTED_POSITION_ROWS:
        fail(f"Lists!AN2:AO19 mismatch.\n  expected {EXPECTED_POSITION_ROWS}\n  actual   {actual_map}")

    sc = wb["Salary Cap"]
    blocks = []  # (posKey, [(attr, primary6, subMin6_or_None)])
    for pos, start_row in EXPECTED_POSITION_ROWS:
        label = sc.cell(row=start_row - 1, column=1).value
        if label != pos:
            fail(f"'Salary Cap'!A{start_row - 1}: expected block label {pos!r}, found {label!r}")
        attrs = []
        r = start_row
        while True:
            attr = sc.cell(row=r, column=2).value
            if attr is None or sc.cell(row=r, column=1).value is not None:
                break
            primary = tuple(sc.cell(row=r, column=c).value for c in range(3, 9))    # C–H
            if not all(is_number(x) for x in primary):
                fail(f"'Salary Cap'!C{r}:H{r} non-numeric primary params for {pos}/{attr}: {primary}")
            sub_raw = tuple(sc.cell(row=r, column=c).value for c in range(9, 15))   # I–N
            sub = sub_raw if all(is_number(x) for x in sub_raw) else None
            attrs.append((attr, primary, sub))
            r += 1
        names = [a for a, _, _ in attrs]
        if names not in (HITTER_ATTRS, PITCHER_ATTRS):
            fail(f"{pos} block attr sequence unexpected: {names}")
        blocks.append((pos, attrs))

    # §3.3 anchor check (STOP on mismatch — never "correct" the spec silently)
    c_attrs = {a: p for a, p, _ in blocks[0][1]}
    for attr, expected in C_ANCHORS.items():
        got = tuple((int(x) if isinstance(x, float) and x.is_integer() else x) for x in c_attrs[attr])
        exp = tuple(expected)
        if got != exp:
            fail(f"C-position anchor mismatch for {attr}: spec {exp} vs workbook {got}")

    # Sub-min curve audit: record exactly which rows carry I–N params
    submin_rows = [(pos, a) for pos, attrs in blocks for a, _, s in attrs if s is not None]
    print(f"[curves] 18 blocks extracted; sub-min reverse params present on: {submin_rows}")
    velo_label = sc.cell(row=79, column=9).value
    if velo_label != "Below Midpoint Velo":
        fail(f"'Salary Cap'!I79 expected 'Below Midpoint Velo' header, found {velo_label!r}")
    return blocks


def read_attr_row(ws, r, deltas_blank_to=0):
    """Read one Traits-sheet pricing row: deltas C–J, flat K, multipliers L–S."""
    deltas, blank_deltas = [], []
    for i, c in enumerate(range(3, 11)):       # C–J → POW..ACC
        v = ws.cell(row=r, column=c).value
        if v is None:
            blank_deltas.append(ALL_ATTRS[i])
            v = deltas_blank_to
        elif not is_number(v):
            fail(f"Traits!{r} col {c}: non-numeric delta {v!r}")
        deltas.append(v)
    flat = ws.cell(row=r, column=11).value      # K
    flat = 0 if flat is None else flat
    mults, blank_mults = [], []
    for i, c in enumerate(range(12, 20)):       # L–S → POW..ACC multipliers
        v = ws.cell(row=r, column=c).value
        if v is None:
            blank_mults.append(ALL_ATTRS[i])
            v = 1
        elif not is_number(v):
            fail(f"Traits!{r} col {c}: non-numeric multiplier {v!r}")
        mults.append(v)
    return deltas, flat, mults, blank_deltas, blank_mults


def extract_traits(wb):
    ws = wb["Traits"]
    # Locate section boundaries by col-A markers
    markers = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v in ("HANDED", "PITCHES", "2nd POSITION", "Arm Angle", "POSITION"):
            markers[v] = r
    for m in ("HANDED", "PITCHES", "2nd POSITION", "Arm Angle", "POSITION"):
        if m not in markers:
            fail(f"Traits sheet: section marker {m!r} not found in column A")

    traits = []
    blanks_report = []
    for r in range(3, markers["HANDED"]):
        label = ws.cell(row=r, column=1).value
        if label is None:
            continue
        chem = ws.cell(row=r, column=2).value
        polarity_cell = ws.cell(row=r, column=21).value  # U: 'Good'/'Bad'
        if polarity_cell not in ("Good", "Bad"):
            fail(f"Traits!U{r}: expected Good/Bad, found {polarity_cell!r}")
        polarity = "positive" if polarity_cell == "Good" else "negative"
        suffix = "(+)" if polarity == "positive" else "(-)"
        if not label.endswith(suffix):
            fail(f"Traits!{r}: name {label!r} suffix disagrees with POS/NEG column {polarity_cell!r}")
        deltas, flat, mults, bd, bm = read_attr_row(ws, r)
        for a in bd:
            blanks_report.append((label, a, "delta", "EMPTY -> 0"))
        traits.append({
            "label": label, "name": label[: -len(suffix)].rstrip(), "chemistry": chem,
            "polarity": polarity, "deltas": deltas, "multipliers": mults, "flat": flat,
            "row": r,
        })
    if len(traits) != 75:
        fail(f"expected 75 traits, extracted {len(traits)}")

    # Flagged-blank verification (§3.6 / §15.3)
    by_label = {t["label"]: t for t in traits}
    print("[traits] §3.6 flagged-blank resolutions:")
    for label, attr in FLAGGED_BLANKS:
        t = by_label[label]
        i = ALL_ATTRS.index(attr)
        cell_col = chr(ord("C") + i)
        raw = ws.cell(row=t["row"], column=3 + i).value
        mult = t["multipliers"][i]
        note = f"delta cell {cell_col}{t['row']} = {'EMPTY -> 0' if raw is None else raw}"
        if mult != 1:
            note += f"; NOTE multiplier col {attr} = {mult} (missed by earlier extraction)"
        print(f"  {label} {attr}: {note}")

    # HANDED block: expect single 'S' row
    handed_rows = []
    for r in range(markers["HANDED"] + 1, markers["PITCHES"]):
        lbl = ws.cell(row=r, column=1).value
        if lbl is not None:
            handed_rows.append((lbl, read_attr_row(ws, r)))
    if [l for l, _ in handed_rows] != ["S"]:
        fail(f"HANDED block: expected ['S'], found {[l for l, _ in handed_rows]}")

    # PITCHES block
    pitch_rows = []
    for r in range(markers["PITCHES"] + 1, markers["2nd POSITION"]):
        lbl = ws.cell(row=r, column=1).value
        if lbl is not None:
            row = read_attr_row(ws, r)
            if any(d != 0 for d in row[0]):
                fail(f"PITCHES row {lbl!r}: expected all-zero deltas, found {row[0]}")
            pitch_rows.append((lbl, row))
    if [l for l, _ in pitch_rows] != ["4F", "2F", "CF", "SL", "CB", "SB", "CH", "FK"]:
        fail(f"PITCHES block labels unexpected: {[l for l, _ in pitch_rows]}")

    # 2nd POSITION block
    secpos_rows = []
    for r in range(markers["2nd POSITION"] + 1, markers["Arm Angle"]):
        lbl = ws.cell(row=r, column=1).value
        if lbl is not None:
            secpos_rows.append((lbl, read_attr_row(ws, r)))
    expected_secpos = ["C", "1B", "2B", "SS", "3B", "LF", "CF", "RF", "IF", "OF", "IF/OF", "1B/OF"]
    if [l for l, _ in secpos_rows] != expected_secpos:
        fail(f"2nd POSITION labels unexpected: {[l for l, _ in secpos_rows]}")

    # Arm Angle block
    arm_rows = []
    for r in range(markers["Arm Angle"] + 1, markers["POSITION"]):
        lbl = ws.cell(row=r, column=1).value
        if lbl is not None:
            arm_rows.append((lbl, read_attr_row(ws, r)))
    if [l for l, _ in arm_rows] != ["High", "Mid", "Low", "Sub"]:
        fail(f"Arm Angle labels unexpected: {[l for l, _ in arm_rows]}")

    return traits, handed_rows, pitch_rows, secpos_rows, arm_rows


def extract_arsenal_tax(wb):
    ws = wb["LeagueSettings"]
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Bullpen Arsenal Costs":
            header_row = r
            break
    if header_row is None:
        fail("LeagueSettings: 'Bullpen Arsenal Costs' header not found")
    if (ws.cell(row=header_row + 1, column=1).value, ws.cell(row=header_row + 1, column=2).value) != ("Pitches", "Cost"):
        fail("LeagueSettings: arsenal table column headers not 'Pitches'/'Cost'")
    table = []
    r = header_row + 2
    while is_number(ws.cell(row=r, column=1).value):
        table.append((int(ws.cell(row=r, column=1).value), ws.cell(row=r, column=2).value))
        r += 1
    if [p for p, _ in table] != list(range(8, 21)):
        fail(f"arsenal tax pitch counts unexpected: {[p for p, _ in table]}")
    return table


def curve_ts(params, indent):
    m, c1, mid, ms, c2, s100 = params
    return (f"{{ min: {num(m)}, curve1: {num(c1)}, mid: {num(mid)}, "
            f"midSal: {num(ms)}, curve2: {num(c2)}, sal100: {num(s100)} }}")


def attr_record_ts(values, indent):
    parts = [f"{a}: {num(v)}" for a, v in zip(ALL_ATTRS, values)]
    return "{ " + ", ".join(parts) + " }"


def write_curves_file(blocks):
    lines = []
    a = lines.append
    a("/**")
    a(" * ivCurves.ts — IV Engine salary-curve parameter table (DATA ONLY, generated).")
    a(" *")
    a(f" * Source workbook: {WORKBOOK_REL} (sheet: 'Salary Cap', cols A:N)")
    a(f" * Extracted: {TODAY} by scripts/extract-iv-data.py — DO NOT EDIT BY HAND; rerun the script.")
    a(f" * Spec: {SPEC_REL} §3.2 (AttributeCurve), §3.3 (position rows), §3.4 (sub-min reverse curve)")
    a(" *")
    a(" * POSITION_ROW_MAP (workbook Lists!AN2:AO19 — first attribute row of each block;")
    a(" * the block's position label sits one row above):")
    a(" *   C->5, 1B->11, 2B->17, SS->23, 3B->29, LF->35, CF->41, RF->47, IF->53, OF->59,")
    a(" *   IF/OF->65, '-'->71, SP->77, SP/RP->85, RP->93, CP->101, 1B/OF->109, EXTRA->117")
    a(" *")
    a(" * Block shapes as found in the workbook:")
    a(" *   - 13 hitter-shaped blocks (POW/CON/SPD/FLD/ARM): C 1B 2B SS 3B LF CF RF IF OF IF/OF '-' 1B/OF")
    a(" *   - 5 pitcher-shaped blocks (POW/CON/SPD/FLD/VEL/JNK/ACC): SP SP/RP RP CP EXTRA")
    a(" *   - Sub-minimum reverse params (cols I-N, 'Below Midpoint Velo') exist ONLY on the")
    a(" *     VEL rows of SP, SP/RP, RP, CP. No other attribute carries them — recorded as-is.")
    a(" */")
    a("")
    a("// Spec §3.2 — verbatim interface shape")
    a("export interface AttributeCurve {")
    a("  min: number;      // rating floor where cost begins")
    a("  curve1: number;   // exponent, segment 1 (min->mid)")
    a("  mid: number;      // rating where segments meet")
    a("  midSal: number;   // $ at mid")
    a("  curve2: number;   // exponent, segment 2 (mid->100)")
    a("  sal100: number;   // $ at rating 100")
    a("}")
    a("")
    a("export type IVAttr = 'POW' | 'CON' | 'SPD' | 'FLD' | 'ARM' | 'VEL' | 'JNK' | 'ACC';")
    a("")
    a("export type PositionKey =")
    a("  | 'C' | '1B' | '2B' | 'SS' | '3B' | 'LF' | 'CF' | 'RF'")
    a("  | 'IF' | 'OF' | 'IF/OF' | '-' | 'SP' | 'SP/RP' | 'RP' | 'CP' | '1B/OF' | 'EXTRA';")
    a("")
    a("export interface AttributeCurveEntry {")
    a("  primary: AttributeCurve;          // workbook cols C-H")
    a("  /** Sub-minimum reverse curve, workbook cols I-N (spec §3.4). Present only where the")
    a("   *  workbook defines it: the VEL row of SP, SP/RP, RP, CP. */")
    a("  subMin?: AttributeCurve;")
    a("}")
    a("")
    a("export interface PositionCurveBlock {")
    a("  // Partial: hitter blocks carry 5 attrs (no VEL/JNK/ACC), pitcher blocks 7 (no ARM)")
    a("  attributes: Partial<Record<IVAttr, AttributeCurveEntry>>;")
    a("}")
    a("")
    a("export const IV_CURVES: Record<PositionKey, PositionCurveBlock> = {")
    for pos, attrs in blocks:
        key = pos if re.fullmatch(r"[A-Za-z_$][A-Za-z0-9_$]*", pos) else f"'{pos}'"
        a(f"  {key}: {{")
        a("    attributes: {")
        for attr, primary, sub in attrs:
            if sub is None:
                a(f"      {attr}: {{ primary: {curve_ts(primary, 0)} }},")
            else:
                a(f"      {attr}: {{")
                a(f"        primary: {curve_ts(primary, 0)},")
                a(f"        subMin: {curve_ts(sub, 0)},")
                a("      },")
        a("    },")
        a("  },")
    a("};")
    a("")
    with open(OUT_CURVES, "w") as f:
        f.write("\n".join(lines))
    print(f"[write] {os.path.relpath(OUT_CURVES, ROOT)} ({len(lines)} lines)")


def write_traits_file(traits, handed, pitches, secpos, arm, arsenal):
    lines = []
    a = lines.append
    a("/**")
    a(" * traitPricing.ts — Trait rating-equivalents + pitch/arsenal/aux pricing (DATA ONLY, generated).")
    a(" *")
    a(f" * Source workbook: {WORKBOOK_REL}")
    a(" *   - 'Traits' sheet rows 3-77: 75 traits (deltas cols C-J, flat fee col K, multipliers cols L-S,")
    a(" *     polarity col U). Values are Chemistry Level 2 baseline (LeagueSettings 'Restrict to Level 2")
    a(" *     Chemistry' = True); potency scaling 0.5/1.0/2.0 applied downstream (spec §3.5).")
    a(" *   - 'Traits' sheet HANDED / PITCHES / 2nd POSITION / Arm Angle blocks (rows 78-106): aux pricing.")
    a(" *   - 'LeagueSettings' A47:B59: bullpen arsenal tax table.")
    a(f" * Extracted: {TODAY} by scripts/extract-iv-data.py — DO NOT EDIT BY HAND; rerun the script.")
    a(f" * Spec: {SPEC_REL} §3.5 (marginal pricing incl. multiplier terms), §3.6 (table + blanks)")
    a(" *")
    a(" * Blank source cells: delta blanks -> 0, multiplier blanks -> 1 (both no-ops in §3.5 pricing).")
    a(" * The §3.6 '·' cells were all verified EMPTY in the workbook, but several of those traits carry")
    a(" * their pricing in the MULTIPLIER columns the §3.6 table omits (e.g. Elite 4F VEL x1.9,")
    a(" * Rally Stopper VEL x1.15, Reverse Splits x1.45/1.4/1.4, Specialist x1.3/1.4/1.3).")
    a(" */")
    a("")
    a("export type PricedAttr = 'POW' | 'CON' | 'SPD' | 'FLD' | 'ARM' | 'VEL' | 'JNK' | 'ACC';")
    a("")
    a("export type ChemistryType = 'Competitive' | 'Crafty' | 'Disciplined' | 'Scholarly' | 'Spirited';")
    a("")
    a("export interface TraitPricingEntry {")
    a("  name: string;                              // workbook label minus ' (+)'/' (-)' suffix")
    a("  chemistry: ChemistryType;")
    a("  polarity: 'positive' | 'negative';")
    a("  deltas: Record<PricedAttr, number>;        // rating-equivalents (L2 baseline)")
    a("  multipliers: Record<PricedAttr, number>;   // attrCost x mult - attrCost terms (1 = no-op)")
    a("  flatFee: number;                           // flat $ added")
    a("}")
    a("")
    a("export const TRAIT_PRICING: TraitPricingEntry[] = [")
    for t in traits:
        a("  {")
        a(f"    name: {t['name']!r}, chemistry: {t['chemistry']!r}, polarity: {t['polarity']!r},")
        a(f"    deltas: {attr_record_ts(t['deltas'], 0)},")
        a(f"    multipliers: {attr_record_ts(t['multipliers'], 0)},")
        a(f"    flatFee: {num(t['flat'])},")
        a("  },")
    a("];")
    a("")
    a("export type PitchType = '4F' | '2F' | 'CF' | 'SL' | 'CB' | 'SB' | 'CH' | 'FK';")
    a("")
    a("export interface PitchCost {")
    a("  flatFee: number;")
    a("  multipliers: Record<PricedAttr, number>;   // pitch deltas are all 0 in source; value is flat+mult")
    a("}")
    a("")
    a("// 'Traits' sheet PITCHES block (rows 81-88)")
    a("export const PITCH_COSTS: Record<PitchType, PitchCost> = {")
    for lbl, (deltas, flat, mults, _, _) in pitches:
        a(f"  '{lbl}': {{ flatFee: {num(flat)}, multipliers: {attr_record_ts(mults, 0)} }},")
    a("};")
    a("")
    a("// LeagueSettings!A47:B59 — total pitch count across the bullpen -> $ adjustment")
    a("export const ARSENAL_TAX_TABLE: Record<number, number> = {")
    a("  " + ", ".join(f"{p}: {num(c)}" for p, c in arsenal) + ",")
    a("};")
    a("")
    a("export interface AuxPricingRow {")
    a("  deltas: Record<PricedAttr, number>;")
    a("  multipliers: Record<PricedAttr, number>;")
    a("  flatFee: number;")
    a("}")
    a("")
    a("// 'Traits' sheet HANDED ('S' row), 2nd POSITION and Arm Angle blocks")
    a("export const AUX_PRICING: {")
    a("  switchHitter: AuxPricingRow;")
    a("  secondaryPositions: Record<string, AuxPricingRow>;")
    a("  armAngle: Record<'High' | 'Mid' | 'Low' | 'Sub', AuxPricingRow>;")
    a("} = {")
    s_deltas, s_flat, s_mults, _, _ = handed[0][1]
    a(f"  switchHitter: {{ deltas: {attr_record_ts(s_deltas, 0)}, multipliers: {attr_record_ts(s_mults, 0)}, flatFee: {num(s_flat)} }},")
    a("  secondaryPositions: {")
    for lbl, (deltas, flat, mults, _, _) in secpos:
        a(f"    '{lbl}': {{ deltas: {attr_record_ts(deltas, 0)}, multipliers: {attr_record_ts(mults, 0)}, flatFee: {num(flat)} }},")
    a("  },")
    a("  armAngle: {")
    for lbl, (deltas, flat, mults, _, _) in arm:
        a(f"    {lbl}: {{ deltas: {attr_record_ts(deltas, 0)}, multipliers: {attr_record_ts(mults, 0)}, flatFee: {num(flat)} }},")
    a("  },")
    a("};")
    a("")
    with open(OUT_TRAITS, "w") as f:
        f.write("\n".join(lines))
    print(f"[write] {os.path.relpath(OUT_TRAITS, ROOT)} ({len(lines)} lines)")


def main():
    if not os.path.exists(WORKBOOK):
        fail(f"workbook not found: {WORKBOOK}")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    blocks = extract_curves(wb)
    traits, handed, pitches, secpos, arm = extract_traits(wb)
    arsenal = extract_arsenal_tax(wb)

    hitter_blocks = [p for p, attrs in blocks if [x for x, _, _ in attrs] == HITTER_ATTRS]
    pitcher_blocks = [p for p, attrs in blocks if [x for x, _, _ in attrs] == PITCHER_ATTRS]
    print(f"[counts] blocks={len(blocks)} (hitter-shaped {len(hitter_blocks)}: {hitter_blocks};")
    print(f"         pitcher-shaped {len(pitcher_blocks)}: {pitcher_blocks})")
    print(f"[counts] traits={len(traits)} pitches={len(pitches)} secondaryPositions={len(secpos)} "
          f"armAngles={len(arm)} arsenalRows={len(arsenal)}")

    write_curves_file(blocks)
    write_traits_file(traits, handed, pitches, secpos, arm, arsenal)
    print("OK")


if __name__ == "__main__":
    main()
