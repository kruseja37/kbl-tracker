#!/usr/bin/env python3
"""
analyze-pool.py — T3: Empirical Pool Analysis & Tier Parameter Derivation

Computes the IV distribution of the 440-player stock SMB4 pool and derives every
empirical constant the spec defers to T3 (tier shifts, tier caps, tier-scaled
luxury caps, farm-draft nerf), then runs the EV-flatness verification (§5.3).

Spec: spec-docs/IV_ENGINE_AND_ROSTER_INTELLIGENCE_SPEC.md §3.2-§3.7, §5.1-§5.3, §6.2-§6.3, §7.4, §13 (T3)

Inputs (all committed, read-only):
  - src/data/playerDatabase.ts          the 440-player stock SMB4 pool (20 teams x 22)
  - src/data/ivCurves.ts                T1 curve params (authoritative post-T1)
  - src/data/traitPricing.ts            T1 trait/pitch/aux pricing (authoritative post-T1)
  - spec-docs/reference/Team_Builder_Archetype_Logic_Template.xlsx
        Roster sheet      -> golden-anchor players (cached salaries) - BOOTSTRAP GATE
        Luxury Cap sheet  -> 19 penalty rows (A:F) + 44 cap modifications (AT:BE)

Output:
  - src/data/tierParams.ts  (regenerated deterministically - no timestamps, no randomness)
  - stdout: all analysis tables (R1-R6) consumed by spec-docs/T3_POOL_ANALYSIS.md

THE BOOTSTRAP RULE: the IV math implemented here was decoded from the workbook's
Roster-sheet formulas (see ivTotal() comments for the exact decode). The script
ABORTS before any analysis unless it reproduces every priced Roster-sheet player's
cached salary within +/-$5 (incl. Eovaldi $54,582 and deGrom $71,609).

src/engines/ivEngine.ts does NOT exist yet (T4); nothing here creates it.

Usage: python3 scripts/analyze-pool.py        (requires openpyxl, like T1's extractor)
"""

import os
import re
import sys
import math
import json
import argparse
import hashlib
import subprocess

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "spec-docs", "reference", "Team_Builder_Archetype_Logic_Template.xlsx")
CURVES_TS = os.path.join(ROOT, "src", "data", "ivCurves.ts")
TRAITS_TS = os.path.join(ROOT, "src", "data", "traitPricing.ts")
PLAYERS_TS = os.path.join(ROOT, "src", "data", "playerDatabase.ts")
OUT_TS = os.path.join(ROOT, "src", "data", "tierParams.ts")

ATTRS8 = ["POW", "CON", "SPD", "FLD", "ARM", "VEL", "JNK", "ACC"]
HITTER_POS = ["C", "1B", "2B", "SS", "3B", "LF", "CF", "RF"]
PITCHER_ROLES = ["SP", "SP/RP", "RP", "CP"]
PITCH_ATTRS = ["VEL", "JNK", "ACC"]
BAT_USAGE_ATTRS = ["POW", "CON", "SPD", "FLD"]
PITCHER_NEUTRAL_HITTER_BLOCK = "IF/OF"
TWO_WAY_TRAIT_POS = {"Two Way (C)": "C", "Two Way (IF)": "IF", "Two Way (OF)": "OF"}
TWO_WAY_ARM_BY_TIER = {"L1": 60, "L2": 80, "L3": 99}
SP_RP_START_SHARE = 0.30
SP_RP_FLEX_PREMIUM = 1.12
SPD_USAGE_FLOORS = {
    "SP": {"prFloor": 0.02, "rangeFloor": 0.10},
    "SP/RP": {"prFloor": 0.02, "rangeFloor": 0.08},
    "RP": {"prFloor": 0.02, "rangeFloor": 0.06},
    "CP": {"prFloor": 0.01, "rangeFloor": 0.05},
}
GRADE_ORDER = ["D", "D+", "C-", "C", "C+", "B-", "B", "B+", "A-", "A", "A+", "S"]
GRADE_ORD = {g: i for i, g in enumerate(GRADE_ORDER)}

# DB trait-name typos (playerDatabase.ts) -> workbook/traitPricing names.
# T3 must not modify the DB (read-only constraint); flagged in T3_POOL_ANALYSIS.md for cleanup.
TRAIT_NAME_FIXES = {
    "Clitch": "Clutch",
    "K Neglecter": "K Neglector",
    "Off-speed Hitter": "Off-Speed Hitter",
}

# Registry constants used (spec §12)
STAR_BUDGET_SHARE = 0.33
ROSTER_HEADROOM = 1.15
LUXURY_CAP_PERCENTILE = 0.65
EV_FLATNESS_TOLERANCE = 0.10
N_TEAMS = 20  # contention ladder depth = league size the 440 pool implies (440/22)

# JK ruling 2026-07-15: pitcher POW/CON stay active at their empirical top-four caps,
# but their response is a quadratic KBL override instead of the workbook's linear cliff.
PITCHER_SECONDARY_BATTING_ROWS = {
    ("rotation", "POW"),
    ("rotation", "CON"),
    ("bullpen", "POW"),
    ("bullpen", "CON"),
}
PITCHER_SECONDARY_BATTING_PENALTY_CURVE = 2.0


# §3.9 registry-style inputs. The script stores the derivation inputs, then emits the
# calibrated weights; downstream kblIV never consumes opaque hand-picked role weights.
USAGE_WEIGHT_INPUTS = {
    # startShare * paRatio + phFloor gives POW/CON exposure. SPD adds PR/range floors.
    "SP": {"startShare": 0.25, "paRatio": 0.625, "phFloor": 0.04, **SPD_USAGE_FLOORS["SP"]},
    "SP/RP": {"startShare": 0.18, "paRatio": 0.625, "phFloor": 0.0375, **SPD_USAGE_FLOORS["SP/RP"]},
    "RP": {"startShare": 0.0, "paRatio": 0.625, "phFloor": 0.08, **SPD_USAGE_FLOORS["RP"]},
    "CP": {"startShare": 0.0, "paRatio": 0.625, "phFloor": 0.05, **SPD_USAGE_FLOORS["CP"]},
}


def role_usage_weights(role):
    inp = USAGE_WEIGHT_INPUTS[role]
    bat = inp["startShare"] * inp["paRatio"] + inp["phFloor"]
    spd = min(1.0, bat + inp["prFloor"] + inp["rangeFloor"])
    return {"POW": bat, "CON": bat, "SPD": spd, "FLD": 1.0}


def emit_usage_weights():
    print("\n" + "=" * 100)
    print("D15/V117. kblIV pitcher-batting usage weights (§3.9) - derived from registry inputs")
    print("=" * 100)
    print("| Role  | startShare | paRatio | phFloor | POW/CON weight | PR floor | range floor | SPD weight | FLD weight |")
    print("|-------|------------|---------|---------|----------------|----------|-------------|------------|------------|")
    for role in PITCHER_ROLES:
        inp = USAGE_WEIGHT_INPUTS[role]
        w = role_usage_weights(role)
        print(f"| {role:<5} | {inp['startShare']:>10.3f} | {inp['paRatio']:>7.3f} | {inp['phFloor']:>7.4f} | "
              f"{w['POW']:>14.4f} | {inp['prFloor']:>8.3f} | {inp['rangeFloor']:>11.3f} | "
              f"{w['SPD']:>10.4f} | {w['FLD']:>10.4f} |")
    print("Pitcher-block batting curves remain rawIV-only; kblIV prices pitcher batting on hitter curves.")
    print(f"SP/RP arm interpolation (D16): alpha={SP_RP_START_SHARE:.2f} SP + "
          f"{1 - SP_RP_START_SHARE:.2f} RP, flexPremium={SP_RP_FLEX_PREMIUM:.2f}.")


def fail(msg):
    print(f"\nT3 ABORT: {msg}", file=sys.stderr)
    sys.exit(1)


def roundup(x):
    """Excel ROUNDUP(x, 0): away from zero. 9-decimal pre-round guards float dust
    (e.g. 19250.000000000004 must stay 19250, matching Excel's 15-sig-digit world)."""
    x = round(x, 9)
    return math.ceil(x) if x >= 0 else math.floor(x)


def percentile(sorted_vals, p):
    """Linear-interpolation percentile (numpy default), p in [0,1]."""
    n = len(sorted_vals)
    if n == 1:
        return sorted_vals[0]
    h = (n - 1) * p
    lo = int(math.floor(h))
    hi = min(lo + 1, n - 1)
    return sorted_vals[lo] + (h - lo) * (sorted_vals[hi] - sorted_vals[lo])


def pstats(vals):
    s = sorted(vals)
    n = len(s)
    mean = sum(s) / n
    sd = math.sqrt(sum((v - mean) ** 2 for v in s) / n)
    return {
        "n": n, "mean": mean, "sd": sd, "min": s[0], "max": s[-1],
        "p10": percentile(s, 0.10), "p25": percentile(s, 0.25),
        "median": percentile(s, 0.50), "p75": percentile(s, 0.75),
        "p90": percentile(s, 0.90),
    }


# ---------------------------------------------------------------------------
# 1. Load T1 data files (authoritative; NEVER re-extracted from the workbook)
# ---------------------------------------------------------------------------

def parse_curves():
    src = open(CURVES_TS).read()
    body = src[src.index("export const IV_CURVES"):]
    curves = {}
    block_re = re.compile(r"^  (?:'([^']+)'|([A-Za-z0-9_]+)): \{$", re.M)
    positions = [(m.group(1) or m.group(2), m.start()) for m in block_re.finditer(body)]
    for i, (pos, start) in enumerate(positions):
        end = positions[i + 1][1] if i + 1 < len(positions) else len(body)
        seg = body[start:end]
        attrs = {}
        for am in re.finditer(r"(\w+): \{ primary: \{ ([^}]+) \} \},", seg):
            attrs[am.group(1)] = {"primary": _params(am.group(2)), "subMin": None}
        for am in re.finditer(r"(\w+): \{\n        primary: \{ ([^}]+) \},\n        subMin: \{ ([^}]+) \},", seg):
            attrs[am.group(1)] = {"primary": _params(am.group(2)), "subMin": _params(am.group(3))}
        curves[pos] = attrs
    if len(curves) != 18:
        fail(f"ivCurves.ts: expected 18 position blocks, parsed {len(curves)}: {sorted(curves)}")
    # spec §3.3 anchor
    c_pow = curves["C"]["POW"]["primary"]
    if (c_pow["min"], c_pow["curve1"], c_pow["mid"], c_pow["midSal"], c_pow["curve2"], c_pow["sal100"]) != (0, 1, 50, 8000, 1.5, 56000):
        fail(f"ivCurves.ts C/POW anchor mismatch: {c_pow}")
    for pos in HITTER_POS:
        if set(curves[pos]) != {"POW", "CON", "SPD", "FLD", "ARM"}:
            fail(f"{pos} block not hitter-shaped: {sorted(curves[pos])}")
    for pos in PITCHER_ROLES:
        if set(curves[pos]) != {"POW", "CON", "SPD", "FLD", "VEL", "JNK", "ACC"}:
            fail(f"{pos} block not pitcher-shaped: {sorted(curves[pos])}")
        if curves[pos]["VEL"]["subMin"] is None:
            fail(f"{pos} VEL subMin params missing")
    return curves


def _params(s):
    out = {}
    for kv in s.split(","):
        k, v = kv.split(":")
        out[k.strip()] = float(v)
    return out


def parse_traits():
    src = open(TRAITS_TS).read()
    traits = {}
    for m in re.finditer(
        r"\{\n    name: '([^']+)', chemistry: '\w+', polarity: '(\w+)',\n"
        r"    deltas: \{ ([^}]+) \},\n    multipliers: \{ ([^}]+) \},\n    flatFee: ([\d.]+),", src):
        traits[m.group(1)] = {
            "polarity": m.group(2),
            "deltas": _params(m.group(3)),
            "multipliers": _params(m.group(4)),
            "flatFee": float(m.group(5)),
        }
    if len(traits) != 75:
        fail(f"traitPricing.ts: expected 75 traits, parsed {len(traits)}")
    pitches = {}
    for m in re.finditer(r"'(\w\w)': \{ flatFee: ([\d.]+), multipliers: \{ ([^}]+) \} \},", src):
        pitches[m.group(1)] = {"flatFee": float(m.group(2)), "multipliers": _params(m.group(3))}
    if len(pitches) != 8:
        fail(f"traitPricing.ts: expected 8 pitch types, parsed {len(pitches)}")
    aux_seg = src[src.index("export const AUX_PRICING"):]
    sw = re.search(r"switchHitter: \{ deltas: \{ ([^}]+) \}, multipliers: \{ ([^}]+) \}, flatFee: ([\d.]+) \}", aux_seg)
    switch = {"deltas": _params(sw.group(1)), "flatFee": float(sw.group(3))}
    secpos = {}
    sec_seg = aux_seg[aux_seg.index("secondaryPositions: {"):aux_seg.index("armAngle: {")]
    for m in re.finditer(r"'([\w/]+)': \{ deltas: \{ ([^}]+) \}, multipliers: \{ ([^}]+) \}, flatFee: ([\d.]+) \}", sec_seg):
        secpos[m.group(1)] = {"deltas": _params(m.group(2)), "flatFee": float(m.group(4))}
    if len(secpos) != 12:
        fail(f"traitPricing.ts: expected 12 secondary-position rows, parsed {len(secpos)}")
    ang_seg = aux_seg[aux_seg.index("armAngle: {"):]
    angles = {}
    for m in re.finditer(r"(\w+): \{ deltas: \{ ([^}]+) \}, multipliers: \{ ([^}]+) \}, flatFee: ([\d.]+) \}", ang_seg):
        angles[m.group(1)] = {"deltas": _params(m.group(2)), "multipliers": _params(m.group(3)), "flatFee": float(m.group(4))}
    if set(angles) != {"High", "Mid", "Low", "Sub"}:
        fail(f"traitPricing.ts: arm angle rows unexpected: {sorted(angles)}")
    return traits, pitches, switch, secpos, angles


# ---------------------------------------------------------------------------
# 2. Load the 440-player pool
# ---------------------------------------------------------------------------

def parse_players():
    src = open(PLAYERS_TS).read()
    start = src.index("export const PLAYERS")
    end = src.index("for (const p of ALL_MLB_PLAYERS)", start)
    block = src[start:end]
    entries = re.findall(r"'([a-z0-9-]+)': \{(.*?)\n  \},?\n", block, re.S)
    pool, fa = [], []
    for pid, body in entries:
        def g(key, default=None):
            m = re.search(key + r": '([^']*)'", body)
            return m.group(1) if m else default
        team = g("teamId")
        p = {"id": pid, "name": g("name"), "team": team, "grade": g("overall"),
             "bats": g("bats"), "isPitcher": "isPitcher: true" in body}
        br = re.search(r"batterRatings: \{ power: (\d+), contact: (\d+), speed: (\d+), fielding: (\d+), arm: (\d+) \}", body)
        if br:
            p["bat"] = dict(zip(["POW", "CON", "SPD", "FLD", "ARM"], map(int, br.groups())))
            p["hasBat"] = True
        elif p["isPitcher"]:
            # Legacy guard only. DB1 closed the stock-pitcher batterRatings gap; if a
            # future DB drops one, rawIV keeps the old non-invented zero default and
            # the analysis reports the regression before F1 luxury derivation.
            p["bat"] = {a: 0 for a in ["POW", "CON", "SPD", "FLD", "ARM"]}
            p["hasBat"] = False
        else:
            fail(f"hitter {pid}: batterRatings missing")
        if p["isPitcher"]:
            pr = re.search(r"pitcherRatings: \{ velocity: (\d+), junk: (\d+), accuracy: (\d+) \}", body)
            if not pr:
                fail(f"pitcher {pid}: pitcherRatings missing")
            p["pit"] = dict(zip(["VEL", "JNK", "ACC"], map(int, pr.groups())))
            p["role"] = g("pitcherRole")
            if p["role"] not in PITCHER_ROLES:
                fail(f"pitcher {pid}: unknown pitcherRole {p['role']!r}")
            p["armSlot"] = g("armSlot")
            if p["armSlot"] not in {"High", "Mid", "Low", "Sub"} and team != "free-agent":
                fail(f"pitcher {pid}: unknown or missing armSlot {p['armSlot']!r}")
            am = re.search(r"arsenal: \[([^\]]*)\]", body)
            p["arsenal"] = re.findall(r"'([^']+)'", am.group(1)) if am else []
        else:
            p["pos"] = g("primaryPosition")
            p["armSlot"] = None
            p["arsenal"] = []
        p["secondary"] = g("secondaryPosition")
        traits = [g("trait1"), g("trait2")]
        p["traits"] = [TRAIT_NAME_FIXES.get(t, t) for t in traits if t]
        (fa if team == "free-agent" else pool).append(p)
    if len(pool) != 440:
        fail(f"playerDatabase.ts: expected 440 team-rostered stock players, found {len(pool)} (+{len(fa)} FA)")
    return pool, fa


# ---------------------------------------------------------------------------
# 3. IV engine (T3-local; decoded from the workbook Roster sheet, 2026-06-10)
#
# Decode provenance (all verified against live formulas in this workbook):
#  - Attribute cell (e.g. Roster!P7): ROUNDUP(twoSegment(rating), 0) per §3.2.
#  - VEL cell (Roster!AE7): if rating <= primary.min -> sub-min reverse branch:
#       reflected = 100 - 100*(r - sub.min) / (primary.min - sub.min)
#       cost = ROUNDUP(twoSegment(reflected, subMinParams), 0)
#    NOTE: the workbook divides by (primary.min - sub.min), NOT the (mid2 - min2)
#    written parenthetically in spec §3.4 - spec amendment candidate (doc §A2).
#  - Trait cell (Roster!AS7): ROUNDUP( sum over attrs of EXACT marginal
#       [twoSegment(r+delta) - twoSegment(r)]  (primary curve only; no sub-min)
#       gated: ARM delta only for hitter-shaped blocks, VEL/JNK/ACC deltas only for
#       pitcher-shaped blocks (workbook gates on position row <77 / >72; block shape
#       is the equivalent, position-table-independent form)
#     + sum over attrs of ROUNDED-CELL multiplier terms [cell*mult - cell]
#     + flatFee, 0)
#  - Switch hitter (Roster!L7): ROUNDUP(exact marginals of HANDED 'S' deltas), same gates.
#  - Pitch cells (PitchCalcs!R2:Y2): per pitch ROUNDUP(flat + sum ROUNDED-CELL
#    multiplier terms over VEL/JNK/ACC), summed per player.
#  - Arm angle (Roster!BT23): 0 unless 'Sub' (flat + rounded-cell mult terms). DB1
#    added armSlot, so pool rawIV/kblIV now price Sub slots; High/Mid/Low stay $0.
#  - Bullpen arsenal tax (LeagueSettings table): TEAM-level (sums the whole pen's
#    pitch count) - NOT part of any player's salary cell; excluded from IV here.
#  - Player salary = SUM of the (already-rounded) component cells. Integer.
# ---------------------------------------------------------------------------

def two_segment(r, c):
    span1 = (c["mid"] - c["min"]) ** c["curve1"]
    seg1 = c["midSal"] * max(r - c["min"], 0) ** c["curve1"] / span1
    top = c["sal100"] - c["midSal"] * max(100 - c["min"], 0) ** c["curve1"] / span1
    if c["mid"] >= 100:  # '-' block: Excel hits 0^0/0 -> IFERROR -> 0
        return seg1
    seg2 = top * max(r - c["mid"], 0) ** c["curve2"] / (100 - c["mid"]) ** c["curve2"]
    return seg1 + seg2


def attr_cell(r, entry):
    prim = entry["primary"]
    if r <= prim["min"]:
        sub = entry["subMin"]
        if sub is None:
            return 0
        reflected = 100 - 100 * (r - sub["min"]) / (prim["min"] - sub["min"])
        return roundup(two_segment(reflected, sub))
    return roundup(two_segment(r, prim))


def marginal(r, delta, entry):
    prim = entry["primary"]
    return two_segment(r + delta, prim) - two_segment(r, prim)


def priced_component(deltas, multipliers, flat, ratings, block, cells, delta_block=None):
    """Shared trait/switch/secondary/angle component pricing (workbook AS-column shape).

    delta_block: curve block used for the EXACT delta marginals. Normally the player's
    own block; for NEGATIVE traits on SP/RP players the workbook routes deltas to the
    RP block (Roster!BW18/BW19 helper: =if(right(trait,3)="(-)","RP","SP/RP")) so the
    hyper-convex dual-role curves can't be farmed for outsized refunds. Multiplier
    terms always reference the player's own ROUNDED attribute cells."""
    dblock = delta_block or block
    hitter_shaped = "ARM" in block
    total = flat
    for a in ["POW", "CON", "SPD", "FLD"]:
        if deltas.get(a, 0) != 0:
            total += marginal(ratings[a], deltas[a], dblock[a])
    if hitter_shaped and deltas.get("ARM", 0) != 0:
        total += marginal(ratings["ARM"], deltas["ARM"], dblock["ARM"])
    if not hitter_shaped:
        for a in ["VEL", "JNK", "ACC"]:
            if deltas.get(a, 0) != 0:
                total += marginal(ratings[a], deltas[a], dblock[a])
    if multipliers:
        for a in ATTRS8:
            m = multipliers.get(a, 1)
            if m != 1 and a in cells:
                total += cells[a] * m - cells[a]
    return roundup(total)


class IVEngine:
    def __init__(self, curves, traits, pitches, switch, secpos, angles):
        self.curves, self.traits, self.pitches = curves, traits, pitches
        self.switch, self.secpos, self.angles = switch, secpos, angles

    def compute(self, block_key, ratings, bats, traits, arsenal, secondary=None, angle=None):
        block = self.curves[block_key]
        cells = {a: attr_cell(ratings[a], block[a]) for a in block if a in ratings}
        parts = {"attributes": sum(cells.values()), "handed": 0, "traits": 0,
                 "pitches": 0, "secondary": 0, "angle": 0}
        if bats == "S":
            parts["handed"] = priced_component(self.switch["deltas"], None, self.switch["flatFee"],
                                               ratings, block, cells)
        for t in traits:
            if t not in self.traits:
                fail(f"trait {t!r} not in traitPricing.ts")
            tr = self.traits[t]
            delta_block = self.curves["RP"] if (block_key == "SP/RP" and tr["polarity"] == "negative") else None
            parts["traits"] += priced_component(tr["deltas"], tr["multipliers"], tr["flatFee"],
                                                ratings, block, cells, delta_block=delta_block)
        for code in arsenal:
            pc = self.pitches[code]
            cost = pc["flatFee"]
            for a in ["VEL", "JNK", "ACC"]:
                m = pc["multipliers"].get(a, 1)
                if m != 1 and a in cells:
                    cost += cells[a] * m - cells[a]
            parts["pitches"] += roundup(cost)
        if secondary:
            sp = self.secpos.get(secondary)
            if sp:  # one stock player has secondary 'P' - no aux row, $0 (doc'd)
                parts["secondary"] = priced_component(sp["deltas"], None, sp["flatFee"],
                                                      ratings, block, cells)
        if angle and angle == "Sub":
            an = self.angles["Sub"]
            parts["angle"] = priced_component(an["deltas"], an["multipliers"], an["flatFee"],
                                              ratings, block, cells)
        parts["total"] = sum(parts.values()) - parts["total"] if "total" in parts else \
            parts["attributes"] + parts["handed"] + parts["traits"] + parts["pitches"] + parts["secondary"] + parts["angle"]
        return parts

    def pool_raw_iv(self, p):
        if p["isPitcher"]:
            ratings = dict(p["bat"], **p["pit"])
            ratings.pop("ARM", None)  # pitchers carry no priced ARM (workbook block shape)
            # secondary=None: the workbook prices secondary positions for hitters only
            # (pitcher Roster rows have no slot); no DB pitcher carries one anyway.
            return self.compute(p["role"], ratings, p["bats"], p["traits"], p["arsenal"], angle=p["armSlot"])
        block_key = "1B" if p["pos"] == "DH" else p["pos"]  # all 8 hitter blocks share identical
        # params in the workbook (T1 data), so the DH->1B mapping is value-neutral; doc'd.
        return self.compute(block_key, dict(p["bat"]), p["bats"], p["traits"], [],
                            secondary=p["secondary"])

    def pool_iv(self, p):
        """Downstream pool IV = kblIV. Non-pitchers remain workbook-exact; pitchers use
        the §3.9 usage layer for batting/fielding while preserving raw pitcher pricing."""
        if not p["isPitcher"]:
            return self.pool_raw_iv(p)
        return self.pool_kbl_iv(p)

    def hitter_block_for_pitcher(self, p):
        for t in p["traits"]:
            if t in TWO_WAY_TRAIT_POS:
                return TWO_WAY_TRAIT_POS[t]
        return PITCHER_NEUTRAL_HITTER_BLOCK

    def pitcher_attr_cell_kbl(self, role, attr, rating):
        if role != "SP/RP":
            return attr_cell(rating, self.curves[role][attr])
        sp = attr_cell(rating, self.curves["SP"][attr])
        rp = attr_cell(rating, self.curves["RP"][attr])
        return roundup((SP_RP_START_SHARE * sp + (1.0 - SP_RP_START_SHARE) * rp) * SP_RP_FLEX_PREMIUM)

    def pitcher_attr_marginal_kbl(self, role, attr, rating, delta):
        if role != "SP/RP":
            return marginal(rating, delta, self.curves[role][attr])
        sp = marginal(rating, delta, self.curves["SP"][attr])
        rp = marginal(rating, delta, self.curves["RP"][attr])
        return (SP_RP_START_SHARE * sp + (1.0 - SP_RP_START_SHARE) * rp) * SP_RP_FLEX_PREMIUM

    def pitcher_kbl_cells(self, p):
        pitch_block = self.curves[p["role"]]
        hitter_block = self.curves[self.hitter_block_for_pitcher(p)]
        is_two_way = any(t in TWO_WAY_TRAIT_POS for t in p["traits"])
        pitch_cells = {a: self.pitcher_attr_cell_kbl(p["role"], a, p["pit"][a]) for a in PITCH_ATTRS}
        bat_cells = {a: attr_cell(p["bat"][a], hitter_block[a]) for a in ["POW", "CON", "SPD"]}
        # POW/CON/SPD are batting usage. FLD is full-use pitcher fielding unless the
        # Two Way trait explicitly unlocks position defense at the trait-position curve.
        fielding_block = hitter_block if is_two_way else pitch_block
        fielding_interpolated = (p["role"] == "SP/RP" and not is_two_way)
        bat_cells["FLD"] = (
            self.pitcher_attr_cell_kbl(p["role"], "FLD", p["bat"]["FLD"])
            if fielding_interpolated
            else attr_cell(p["bat"]["FLD"], fielding_block["FLD"])
        )
        return pitch_block, hitter_block, fielding_block, fielding_interpolated, pitch_cells, bat_cells

    def weighted_component(self, deltas, multipliers, flat, ratings, pitch_block, hitter_block,
                           fielding_block, fielding_interpolated, pitch_cells, bat_cells, weights, role):
        """kblIV trait/switch component for pitchers.

        Raw workbook semantics price pitcher batting on pitcher blocks. The kbl layer
        routes POW/CON/SPD through hitter curves with usage weights, keeps FLD as full-use
        mound fielding unless Two Way unlocks position defense, and routes SP/RP pitching
        deltas through the same D16 interpolation used by base cells. The old rawIV A3
        negative-trait asymmetry is intentionally not mirrored here; JK ratified the D16
        symmetric flex-premium treatment. Ordinary pitcher ARM remains unpriced.
        """
        total = flat
        for a in ["POW", "CON", "SPD"]:
            if deltas.get(a, 0) != 0:
                total += marginal(ratings[a], deltas[a], hitter_block[a]) * weights[a]
        if deltas.get("FLD", 0) != 0:
            if fielding_interpolated:
                total += self.pitcher_attr_marginal_kbl(role, "FLD", ratings["FLD"], deltas["FLD"])
            else:
                total += marginal(ratings["FLD"], deltas["FLD"], fielding_block["FLD"])
        for a in PITCH_ATTRS:
            if deltas.get(a, 0) != 0:
                if role == "SP/RP":
                    total += self.pitcher_attr_marginal_kbl(role, a, ratings[a], deltas[a])
                else:
                    total += marginal(ratings[a], deltas[a], pitch_block[a])
        if multipliers:
            for a in ["POW", "CON", "SPD"]:
                m = multipliers.get(a, 1)
                if m != 1:
                    total += (bat_cells[a] * m - bat_cells[a]) * weights[a]
            m = multipliers.get("FLD", 1)
            if m != 1:
                total += bat_cells["FLD"] * m - bat_cells["FLD"]
            for a in PITCH_ATTRS:
                m = multipliers.get(a, 1)
                if m != 1:
                    total += pitch_cells[a] * m - pitch_cells[a]
        return roundup(total)

    def two_way_trait_component(self, trait_name, p, hitter_block, bat_cells, weights):
        """§3.9 Two Way as usage unlock, not the workbook's flat +rating delta row.

        Base pitcher kblIV already prices role-expected batting use and full pitcher
        fielding. The Two Way trait buys the remaining everyday hitting usage plus
        quality-scaled defensive capability at the trait position: L2 FLD delta from the
        trait row and L2 ARM = 80 on the trait-position ARM curve.
        """
        tr = self.traits[trait_name]
        total = 0.0
        for a in ["POW", "CON", "SPD"]:
            total += bat_cells[a] * (1.0 - weights[a])
        fld_delta = tr["deltas"].get("FLD", 0)
        if fld_delta:
            total += marginal(p["bat"]["FLD"], fld_delta, hitter_block["FLD"])
        total += attr_cell(TWO_WAY_ARM_BY_TIER["L2"], hitter_block["ARM"])
        return roundup(total)

    def pool_kbl_iv(self, p):
        pitch_block, hitter_block, fielding_block, fielding_interpolated, pitch_cells, bat_cells = self.pitcher_kbl_cells(p)
        weights = role_usage_weights(p["role"])
        parts = {
            "pitchingAttributes": sum(pitch_cells.values()),
            "battingAttributes": roundup(sum(bat_cells[a] * weights[a] for a in BAT_USAGE_ATTRS)),
            "attributes": 0,
            "handed": 0,
            "traits": 0,
            "twoWayUnlock": 0,
            "pitches": 0,
            "secondary": 0,
            "angle": 0,
        }
        parts["attributes"] = parts["pitchingAttributes"] + parts["battingAttributes"]
        ratings = dict(p["bat"], **p["pit"])
        if p["bats"] == "S":
            parts["handed"] = self.weighted_component(self.switch["deltas"], None, self.switch["flatFee"],
                                                      ratings, pitch_block, hitter_block, fielding_block,
                                                      fielding_interpolated, pitch_cells, bat_cells, weights, p["role"])
        for t in p["traits"]:
            if t not in self.traits:
                fail(f"trait {t!r} not in traitPricing.ts")
            if t in TWO_WAY_TRAIT_POS:
                v = self.two_way_trait_component(t, p, hitter_block, bat_cells, weights)
                parts["twoWayUnlock"] += v
                parts["traits"] += v
                continue
            tr = self.traits[t]
            parts["traits"] += self.weighted_component(tr["deltas"], tr["multipliers"], tr["flatFee"],
                                                       ratings, pitch_block, hitter_block, fielding_block,
                                                       fielding_interpolated, pitch_cells, bat_cells, weights, p["role"])
        for code in p["arsenal"]:
            pc = self.pitches[code]
            cost = pc["flatFee"]
            for a in PITCH_ATTRS:
                m = pc["multipliers"].get(a, 1)
                if m != 1:
                    cost += pitch_cells[a] * m - pitch_cells[a]
            parts["pitches"] += roundup(cost)
        if p["armSlot"] == "Sub":
            an = self.angles["Sub"]
            cost = an["flatFee"]
            for a in PITCH_ATTRS:
                m = an["multipliers"].get(a, 1)
                if m != 1:
                    cost += pitch_cells[a] * m - pitch_cells[a]
            parts["angle"] = roundup(cost)
        parts["total"] = parts["attributes"] + parts["handed"] + parts["traits"] + parts["pitches"] + parts["angle"]
        parts["usageWeights"] = weights
        parts["hitterCurveBlock"] = next(k for k, v in self.curves.items() if v is hitter_block)
        parts["fieldingCurveBlock"] = "SP/RP interpolated" if fielding_interpolated else (
            parts["hitterCurveBlock"] if any(t in TWO_WAY_TRAIT_POS for t in p["traits"]) else p["role"]
        )
        return parts


# ---------------------------------------------------------------------------
# 4. Golden anchors - live from the workbook Roster sheet (BOOTSTRAP GATE)
# ---------------------------------------------------------------------------

def load_anchors():
    try:
        import openpyxl
    except ImportError:
        fail("openpyxl required (same dependency as T1's extract-iv-data.py)")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb["Roster"]
    pcols = [("4F", 64), ("2F", 65), ("CF", 66), ("CH", 67), ("CB", 68), ("SL", 69), ("FK", 70), ("SB", 71)]
    anchors = []
    for row in range(6, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value
        pos = ws.cell(row=row, column=8).value
        sal = ws.cell(row=row, column=60).value
        if not isinstance(name, str) or pos not in (HITTER_POS + PITCHER_ROLES) or not isinstance(sal, (int, float)):
            continue
        ratings = {}
        for a, col in [("POW", 16), ("CON", 19), ("SPD", 22), ("FLD", 25), ("ARM", 28),
                       ("VEL", 31), ("JNK", 34), ("ACC", 37)]:
            v = ws.cell(row=row, column=col).value
            if isinstance(v, (int, float)):
                ratings[a] = float(v)
        traits = []
        for col in (45, 52):
            t = ws.cell(row=row, column=col).value
            if isinstance(t, str) and t != "--":
                traits.append(re.sub(r" \([+-]\)$", "", t))
        arsenal = [code for code, col in pcols if ws.cell(row=row, column=col).value is True]
        sec = ws.cell(row=row, column=10).value  # Roster col J = secondary position ('-' = none)
        anchors.append({
            "name": name, "block": pos, "ratings": ratings,
            "bats": ws.cell(row=row, column=12).value,
            "traits": traits, "arsenal": arsenal,
            "secondary": sec if isinstance(sec, str) and sec != "-" else None,
            "angle": ws.cell(row=row, column=72).value,
            "expected": int(sal),
        })
    if len(anchors) < 4:
        fail(f"Roster sheet: found only {len(anchors)} priced players (need >=4)")
    return anchors


def run_anchor_gate(engine, anchors):
    print("=" * 100)
    print("GOLDEN ANCHOR GATE - workbook Roster sheet cached salaries vs T3 IV engine (tolerance +/-$5)")
    print("=" * 100)
    rows, worst = [], None
    for a in anchors:
        parts = engine.compute(a["block"], a["ratings"], a["bats"], a["traits"], a["arsenal"],
                               secondary=a["secondary"], angle=a["angle"])
        diff = parts["total"] - a["expected"]
        rows.append((a["name"], a["block"], a["expected"], parts["total"], diff, parts))
        if worst is None or abs(diff) > abs(worst[4]):
            worst = rows[-1]
    print(f"| {'Player':<18} | {'Pos':<5} | {'Workbook $':>10} | {'Computed $':>10} | {'Diff':>5} | Verdict |")
    print("|" + "-" * 20 + "|" + "-" * 7 + "|" + "-" * 12 + "|" + "-" * 12 + "|" + "-" * 7 + "|---------|")
    n_pass = 0
    for name, blk, exp, got, diff, _ in rows:
        ok = abs(diff) <= 5
        n_pass += ok
        print(f"| {name:<18} | {blk:<5} | {exp:>10,} | {got:>10,} | {diff:>+5} | {'PASS' if ok else 'FAIL'}    |")
    must = {"Nathan Eovaldi": 54582, "Jacob deGROM": 71609}
    seen = {r[0]: r[2] for r in rows}
    for nm, expected in must.items():
        if seen.get(nm) != expected:
            fail(f"contract anchor {nm} (${expected:,}) not found in Roster sheet as expected (saw {seen.get(nm)})")
    gray = next((a for a in anchors if a["name"] == "Jon Gray"), None)
    if gray is None:
        fail("Jon Gray rawIV Injury Prone anchor not found")
    block = engine.curves["SP/RP"]
    cells = {a: attr_cell(gray["ratings"][a], block[a]) for a in block if a in gray["ratings"]}
    injury = engine.traits["Injury Prone"]
    gray_injury = priced_component(injury["deltas"], injury["multipliers"], injury["flatFee"],
                                   gray["ratings"], block, cells, delta_block=engine.curves["RP"])
    if gray_injury != -2136:
        fail(f"Jon Gray rawIV Injury Prone anchor moved: expected -$2,136, saw {gray_injury}")
    if n_pass != len(rows):
        name, blk, exp, got, diff, parts = worst
        print(f"\nWorst mismatch: {name} ({blk}) expected {exp:,} got {got:,} (diff {diff:+})")
        print("Per-component breakdown:", {k: v for k, v in parts.items()})
        fail("golden anchors FAILED - IV math is broken; no analysis was run (bootstrap rule)")
    print(f"\nANCHOR GATE: {n_pass}/{len(rows)} PASS (incl. Eovaldi $54,582, deGrom $71,609; Jon Gray Injury Prone -$2,136)\n")
    return len(rows)


# ---------------------------------------------------------------------------
# T4 oracle dump (serialization-only; reuses the engine objects above)
# ---------------------------------------------------------------------------

def _git_sha():
    try:
        return subprocess.check_output(["git", "rev-parse", "HEAD"], cwd=ROOT, text=True).strip()
    except Exception:
        return "unknown"


def _sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _parts_json(parts):
    out = {}
    for k in sorted(parts):
        v = parts[k]
        if isinstance(v, dict):
            out[k] = {kk: v[kk] for kk in sorted(v)}
        else:
            out[k] = v
    return out


def _player_profile(p):
    return {
        "id": p["id"],
        "name": p["name"],
        "team": p["team"],
        "grade": p["grade"],
        "isPitcher": p["isPitcher"],
        "position": p["role"] if p["isPitcher"] else p["pos"],
        "role": p["role"] if p["isPitcher"] else None,
        "primaryPosition": None if p["isPitcher"] else p["pos"],
        "secondaryPosition": p.get("secondary"),
        "bats": p["bats"],
        "traits": list(p["traits"]),
        "arsenal": list(p["arsenal"]),
        "armSlot": p.get("armSlot"),
        "batterRatings": dict(p["bat"]),
        "pitcherRatings": dict(p["pit"]) if p["isPitcher"] else None,
    }


def _anchor_profile(a):
    return {
        "name": a["name"],
        "position": a["block"],
        "role": a["block"] if a["block"] in PITCHER_ROLES else None,
        "primaryPosition": None if a["block"] in PITCHER_ROLES else a["block"],
        "ratings": dict(a["ratings"]),
        "bats": a["bats"],
        "traits": list(a["traits"]),
        "arsenal": list(a["arsenal"]),
        "secondaryPosition": a["secondary"],
        "armSlot": a["angle"],
    }


def build_iv_oracle(engine, anchors, pool, anchor_count):
    gray = next((a for a in anchors if a["name"] == "Jon Gray"), None)
    if gray is None:
        fail("Jon Gray rawIV Injury Prone anchor not found for oracle serialization")
    block = engine.curves["SP/RP"]
    cells = {a: attr_cell(gray["ratings"][a], block[a]) for a in block if a in gray["ratings"]}
    injury = engine.traits["Injury Prone"]
    gray_injury = priced_component(injury["deltas"], injury["multipliers"], injury["flatFee"],
                                   gray["ratings"], block, cells, delta_block=engine.curves["RP"])
    return {
        "meta": {
            "gitSha": _git_sha(),
            "analyzePoolSha256": _sha256_file(os.path.abspath(__file__)),
            "anchorGate": {"passed": True, "count": anchor_count, "jonGrayInjuryProneDelta": gray_injury},
        },
        "anchors": [
            {
                "name": a["name"],
                "expected": a["expected"],
                "computedRawIV": engine.compute(a["block"], a["ratings"], a["bats"], a["traits"], a["arsenal"],
                                                secondary=a["secondary"], angle=a["angle"])["total"],
                "input": _anchor_profile(a),
            }
            for a in anchors
        ],
        "players": [
            {
                "id": p["id"],
                "name": p["name"],
                "position": p["role"] if p["isPitcher"] else p["pos"],
                "role": p["role"] if p["isPitcher"] else None,
                "rawIV": p["rawIV"],
                "kblIV": p["iv"],
                "rawComponents": _parts_json(p["rawParts"]),
                "kblComponents": _parts_json(p["parts"]),
                "input": _player_profile(p),
            }
            for p in sorted(pool, key=lambda x: x["id"])
        ],
    }


def write_iv_oracle(path, oracle):
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w") as f:
        json.dump(oracle, f, indent=2, sort_keys=True)
        f.write("\n")
    print(f"[write] IV oracle: {os.path.relpath(path, ROOT)} ({len(oracle['players'])} players)")


# ---------------------------------------------------------------------------
# 5. Workbook Luxury Cap sheet - penalty rows + 44 modifications (live read)
# ---------------------------------------------------------------------------

def load_luxury(wbpath):
    import openpyxl
    wb = openpyxl.load_workbook(wbpath, data_only=True)
    ws = wb["Luxury Cap"]
    groups = [("hitters", 3, 8, ["POW", "CON", "SPD", "FLD", "ARM"]),
              ("rotation", 9, 4, ["POW", "CON", "SPD", "FLD", "VEL", "JNK", "ACC"]),
              ("bullpen", 17, None, ["POW", "CON", "SPD", "FLD", "VEL", "JNK", "ACC"])]
    rows = []
    for gname, hdr_row, _, stats in groups:
        r = hdr_row + 1
        for stat in stats:
            label = ws.cell(row=r, column=1).value
            if label != stat:
                fail(f"Luxury Cap sheet: expected {stat} at A{r}, found {label!r}")
            rows.append({
                "group": gname, "stat": stat,
                "curve": float(ws.cell(row=r, column=2).value),
                "topN": int(ws.cell(row=r, column=3).value),
                "xblCap": float(ws.cell(row=r, column=4).value),
                "per100": float(ws.cell(row=r, column=5).value),
                "minAdder": float(ws.cell(row=r, column=6).value),
            })
            r += 1
    if len(rows) != 19:
        fail(f"Luxury Cap sheet: expected 19 penalty rows, got {len(rows)}")
    for r in rows:
        r["enabled"] = True
    spec_pins = {("hitters", "POW"): (1.5, 8, 500, 1500000, 3000),
                 ("bullpen", "VEL"): (1.1, 3, 65, 3000000, 5000)}
    for (g, s), (cv, n, cap, p100, ma) in spec_pins.items():
        row = next(r for r in rows if r["group"] == g and r["stat"] == s)
        if (row["curve"], row["topN"], row["xblCap"], row["per100"], row["minAdder"]) != (cv, n, cap, p100, ma):
            fail(f"Luxury Cap row {g}/{s} disagrees with spec §5.3: {row}")
    # Preserve the workbook as the raw source, then apply the approved KBL product override.
    # This happens before R4/R5 so analysis, generated data, and runtime all use one curve.
    for row in rows:
        if (row["group"], row["stat"]) in PITCHER_SECONDARY_BATTING_ROWS:
            row["curve"] = PITCHER_SECONDARY_BATTING_PENALTY_CURVE

    # 44 modifications, AT:BE = cols 46 (name) + 47-57 (11 deltas); header at row 2
    mod_stats = ["POW", "CON", "SPD", "FLD", "ARM", "RVEL", "RJNK", "RACC", "PVEL", "PJNK", "PACC"]
    hdr = [ws.cell(row=2, column=c).value for c in range(46, 58)]
    if hdr != ["Modification", "POW", "CON", "SPD", "FLD", "ARM", "ROT VEL", "ROT JNK", "ROT ACC",
               "PEN VEL", "PEN JNK", "PEN ACC"]:
        fail(f"Luxury Cap AT:BE header row unexpected: {hdr}")
    mods = {}
    r = 3
    while True:
        name = ws.cell(row=r, column=46).value
        if name is None:
            break
        deltas = [ws.cell(row=r, column=47 + i).value for i in range(11)]
        if any(not isinstance(d, (int, float)) for d in deltas):
            fail(f"Luxury Cap mod row {r} ({name!r}): non-numeric deltas {deltas}")
        mods[name] = dict(zip(mod_stats, [float(d) for d in deltas]))
        r += 1
    # Spec prose says "44 modifications" but the workbook AT:BE block AND the spec's own
    # §6.2 table both contain 42 rows (41 named + neutral '--'). Workbook is source of
    # truth; the count discrepancy is a spec amendment candidate (see T3_POOL_ANALYSIS.md).
    if len(mods) != 42:
        fail(f"Luxury Cap sheet: expected 42 modifications (incl '--'), got {len(mods)}")
    if mods["Defense First"]["FLD"] != 337 or mods["Call Your Shot"]["POW"] != 50:
        fail("Luxury Cap mods disagree with spec §6.2 pins (Defense First FLD=337 / Call Your Shot POW=50)")
    return rows, mods, mod_stats


# ---------------------------------------------------------------------------
# R1 - IV distribution
# ---------------------------------------------------------------------------

def stat_row(label, st):
    return (f"| {label:<14} | {st['n']:>3} | {st['mean']:>10,.0f} | {st['median']:>10,.0f} | "
            f"{st['p10']:>9,.0f} | {st['p25']:>9,.0f} | {st['p75']:>10,.0f} | {st['p90']:>10,.0f} | "
            f"{st['max']:>10,.0f} | {st['sd']:>9,.0f} |")


STAT_HDR = ("| Segment        |   n |       mean |     median |       p10 |       p25 |        p75 |        p90 |"
            "        max |        sd |\n"
            "|----------------|-----|------------|------------|-----------|-----------|------------|------------|------------|-----------|")


def r1_distribution(pool):
    print("=" * 100)
    print("R1. kblIV DISTRIBUTION - 440-player stock SMB4 pool (Juiced tier by definition, D3 + §3.9)")
    print("=" * 100)
    ivs = [p["iv"] for p in pool]
    print("\n### R1a. Overall + by role\n")
    print(STAT_HDR)
    print(stat_row("ALL (440)", pstats(ivs)))
    hitters = [p for p in pool if not p["isPitcher"]]
    print(stat_row("Hitters", pstats([p["iv"] for p in hitters])))
    for role in PITCHER_ROLES:
        seg = [p["iv"] for p in pool if p["isPitcher"] and p["role"] == role]
        print(stat_row(role, pstats(seg)))
    print("\n### R1b. Hitters by primary position\n")
    print(STAT_HDR)
    for pos in HITTER_POS + ["DH"]:
        seg = [p["iv"] for p in hitters if p["pos"] == pos]
        if seg:
            print(stat_row(pos, pstats(seg)))
    print("\n### R1c. Grade overlay (SMB4 letter grades carried by the DB)\n")
    print("| Grade | n   | mean IV    | median IV  | min IV     | max IV     |")
    print("|-------|-----|------------|------------|------------|------------|")
    for g in reversed(GRADE_ORDER):
        seg = [p["iv"] for p in pool if p["grade"] == g]
        if seg:
            st = pstats(seg)
            print(f"| {g:<5} | {st['n']:>3} | {st['mean']:>10,.0f} | {st['median']:>10,.0f} | "
                  f"{st['min']:>10,.0f} | {st['max']:>10,.0f} |")
    return ivs


# ---------------------------------------------------------------------------
# R2 - tier shifts (grade-ladder method) + farm nerf
# ---------------------------------------------------------------------------

def grade_ladder(pool):
    med = {}
    for g in GRADE_ORDER:
        seg = sorted(p["iv"] for p in pool if p["grade"] == g)
        if seg:
            med[GRADE_ORD[g]] = percentile(seg, 0.5)
    return med


def ladder_interp(med, x):
    ords = sorted(med)
    x = max(min(x, ords[-1]), ords[0])  # clamp to observed grade range (doc'd)
    for lo, hi in zip(ords, ords[1:]):
        if lo <= x <= hi:
            f = (x - lo) / (hi - lo)
            return med[lo] + f * (med[hi] - med[lo])
    return med[ords[-1]]


def r2_tiers(pool, ivs):
    print("\n" + "=" * 100)
    print("R2. TIER DERIVATION (§5.1) - grade-ladder method, multiplicative IV-space transform")
    print("=" * 100)
    med = grade_ladder(pool)
    mean_ord = sum(GRADE_ORD[p["grade"]] for p in pool) / len(pool)
    mean_iv = sum(ivs) / len(ivs)
    print(f"\nPool mean grade ordinal: {mean_ord:.3f} "
          f"(ladder: {' < '.join(GRADE_ORDER)}; B- = 5, B = 6)")
    print("\n### Grade ladder (median IV per grade, the empirical grade->$ map)\n")
    print("| Grade | ordinal | median IV |")
    print("|-------|---------|-----------|")
    for o in sorted(med, reverse=True):
        print(f"| {GRADE_ORDER[o]:<5} | {o:>7} | {med[o]:>9,.0f} |")
    anchor = ladder_interp(med, mean_ord)
    scales = {}
    for tier, steps in [("juiced", 0), ("standard", 1), ("nerfed", 2)]:
        target = ladder_interp(med, mean_ord - steps)
        scales[tier] = target / anchor
    print(f"\nLadder IV at pool mean ordinal ({mean_ord:.3f}): {anchor:,.0f}")
    print("\n### Tier scale factors (multiplicative: tier IV distribution = Juiced IV x scale)\n")
    print("| Tier     | grade steps left | ladder target IV | scale  | implied mean IV | implied median IV |")
    print("|----------|------------------|------------------|--------|-----------------|-------------------|")
    s_med = percentile(sorted(ivs), 0.5)
    for tier, steps in [("juiced", 0), ("standard", 1), ("nerfed", 2)]:
        s = scales[tier]
        print(f"| {tier:<8} | {steps:>16} | {ladder_interp(med, mean_ord - steps):>16,.0f} | {s:>6.4f} | "
              f"{mean_iv * s:>15,.0f} | {s_med * s:>17,.0f} |")
    print("""
Transform choice (justification, doc'd in full in T3_POOL_ANALYSIS.md):
  multiplicative scale (IV' = s x IV) over additive shift / quantile remap because
  (a) IV is bounded below by ~$0 and grade steps are multiplicative in $ (convex curves),
  (b) it preserves the observed distribution SHAPE (a 'bell shifted left' in log-$),
  (c) the Player Generator can apply it directly (scale target quantiles by s),
  (d) zero free parameters beyond the empirically derived scale.""")
    # cross-check: adjacent-grade step ratios around the mean
    lo, hi = int(math.floor(mean_ord)), int(math.ceil(mean_ord))
    print("Cross-check - adjacent grade-step ratios near the pool mean: " +
          ", ".join(f"{GRADE_ORDER[o]}->{GRADE_ORDER[o-1]}: {med[o-1]/med[o]:.4f}"
                    for o in (hi, lo) if o - 1 in med and o in med))
    # farm nerf: one grade step left of the league tier (no new free parameter)
    print("\n### Farm-draft nerf (§7.4): one additional grade step left of the league tier\n")
    print("| League tier | farm ordinal target | farmScale (vs tier) | farmScale (vs Juiced) |")
    print("|-------------|---------------------|---------------------|-----------------------|")
    farm = {}
    for tier, steps in [("juiced", 0), ("standard", 1), ("nerfed", 2)]:
        o_tier = mean_ord - steps
        ratio = ladder_interp(med, o_tier - 1) / ladder_interp(med, o_tier)
        farm[tier] = ratio
        print(f"| {tier:<11} | {o_tier - 1:>19.3f} | {ratio:>19.4f} | {scales[tier] * ratio:>21.4f} |")
    # star-rarity evidence
    s_ivs = sorted(ivs)
    print("\nStar-rarity evidence - P(farm player out-IVs the league tier's p75/p90 player):\n")
    print("| League tier | P(farm >= tier p75) | P(farm >= tier p90) |")
    print("|-------------|---------------------|---------------------|")
    for tier in ["juiced", "standard", "nerfed"]:
        r = farm[tier]
        # farm dist = tier dist x r; P(tier_X * r >= tier_q) = P(X >= q/r) on the pool shape
        for label, q in [("p75", percentile(s_ivs, 0.75)), ("p90", percentile(s_ivs, 0.90))]:
            pass
        p75 = sum(1 for v in s_ivs if v * r >= percentile(s_ivs, 0.75)) / len(s_ivs)
        p90 = sum(1 for v in s_ivs if v * r >= percentile(s_ivs, 0.90)) / len(s_ivs)
        print(f"| {tier:<11} | {p75:>19.1%} | {p90:>19.1%} |")
    return scales, farm, med, mean_ord, mean_iv


# ---------------------------------------------------------------------------
# R3 - tier caps (§5.2)
# ---------------------------------------------------------------------------

def r3_caps(ivs, scales):
    print("\n" + "=" * 100)
    print("R3. TIER CAPS (§5.2): tierCap = max(maxIV/0.33, 22 x medianIV x 1.15)")
    print("=" * 100 + "\n")
    s_ivs = sorted(ivs)
    max_iv, med_iv = s_ivs[-1], percentile(s_ivs, 0.5)
    caps = {}
    print("| Tier     | maxIV      | medianIV  | star branch (max/0.33) | roster branch (22*med*1.15) | tierCap    | dominant | ratio |")
    print("|----------|------------|-----------|-------------------------|------------------------------|------------|----------|-------|")
    for tier in ["juiced", "standard", "nerfed"]:
        s = scales[tier]
        b1 = max_iv * s / STAR_BUDGET_SHARE
        b2 = 22 * med_iv * s * ROSTER_HEADROOM
        cap = max(b1, b2)
        ratio = max(b1, b2) / min(b1, b2)
        caps[tier] = {"cap": cap, "starBranch": b1, "rosterBranch": b2,
                      "maxIV": max_iv * s, "medianIV": med_iv * s}
        dom = "star" if b1 > b2 else "roster"
        flag = "  <-- FLAG: star branch dominates >1.5x (starBudgetShare retune signal)" \
            if b1 > 1.5 * b2 else ""
        print(f"| {tier:<8} | {max_iv*s:>10,.0f} | {med_iv*s:>9,.0f} | {b1:>23,.0f} | {b2:>28,.0f} | "
              f"{cap:>10,.0f} | {dom:<8} | {ratio:>5.2f} |{flag}")
    print("\n(Both branches scale linearly with the tier scale s, so the dominant branch and the")
    print(" ratio are tier-invariant under the multiplicative transform - shown per-tier for audit.)")
    return caps


# ---------------------------------------------------------------------------
# R4 - luxury cap scaling (§5.3): contention-ladder percentile method
# ---------------------------------------------------------------------------

def subpool_stat(pool, group, stat):
    if group == "hitters":
        return sorted((p["bat"][stat] for p in pool if not p["isPitcher"]), reverse=True)
    roles = {"rotation": ("SP", "SP/RP"), "bullpen": ("RP", "CP", "SP/RP")}[group]
    sel = [p for p in pool if p["isPitcher"] and p["role"] in roles]
    if stat in ("VEL", "JNK", "ACC"):
        return sorted((p["pit"][stat] for p in sel), reverse=True)
    return sorted((p["bat"][stat] for p in sel), reverse=True)


def contention_ladder(values_desc, top_n, n_teams=N_TEAMS):
    """S(k) = best top-N sum a focused team amasses when k teams contend for this stat
    (every k-th rank from the top). k=1 = degenerate hoard; k=20 = all-contend snake."""
    out = []
    for k in range(1, n_teams + 1):
        picks = [values_desc[j * k] for j in range(top_n) if j * k < len(values_desc)]
        out.append(sum(picks))
    return out


def stock_team_sums(pool, group, stat, top_n):
    """Top-N sum of this stat on each of the 20 REAL stock SMB4 rosters - the pool's
    native 'plausible roster' distribution (the manufacturer's own balanced teams)."""
    teams = sorted({p["team"] for p in pool})
    out = []
    for t in teams:
        members = [p for p in pool if p["team"] == t]
        if group == "hitters":
            vals = [p["bat"][stat] for p in members if not p["isPitcher"]]
        else:
            roles = {"rotation": ("SP", "SP/RP"), "bullpen": ("RP", "CP", "SP/RP")}[group]
            sel = [p for p in members if p["isPitcher"] and p["role"] in roles]
            vals = [p["pit"][stat] if stat in ("VEL", "JNK", "ACC") else p["bat"][stat] for p in sel]
        out.append(sum(sorted(vals, reverse=True)[:top_n]))
    return out


def r4_luxury(pool, lux_rows, mods, mod_stats, scales, anchors_median):
    print("\n" + "=" * 100)
    print(f"R4. LUXURY CAP SCALING (§5.3) - neutral cap = {LUXURY_CAP_PERCENTILE:.0%} percentile of the")
    print("    best-plausible top-N sum distribution (contention ladder S(k), k=1..20 teams)")
    print("=" * 100)
    med_iv = percentile(sorted(p["iv"] for p in pool), 0.5)
    sigma_j = med_iv / anchors_median
    print(f"\nPenalty-$ scale: sigma(juiced) = pool median IV / XBL anchor median salary "
          f"= {med_iv:,.0f} / {anchors_median:,.0f} = {sigma_j:.4f}")
    print("(keeps tax bite per overage point proportional to this pool's salary scale;")
    print(" penalty CURVES port unchanged except the approved quadratic pitcher POW/CON KBL override.)\n")
    derived = []
    print("Two candidate 'best-plausible top-N sum' distributions were derived; the STOCK-TEAM")
    print("basis is ADOPTED (rationale + the rejected alternative documented in T3_POOL_ANALYSIS.md §R4):")
    print("  A. stock-team basis [ADOPTED]: top-N sums of each of the 20 real SMB4 rosters -")
    print("     caps bind just above typical real-team concentration, keeping the tax layer live.")
    print("  B. contention ladder [REJECTED]: S(k) = focused team's take under k-team contention,")
    print("     ignores budget -> caps so high the tax layer never binds (EV test went vacuous).\n")
    print("| Group    | Stat | topN | XBL cap | teams min | teams p50 | 65th pct -> cap (juiced) | teams max | ladder-B cap | vs XBL | status |")
    print("|----------|------|------|---------|-----------|-----------|--------------------------|-----------|--------------|--------|--------|")
    for row in lux_rows:
        vals = subpool_stat(pool, row["group"], row["stat"])
        ladder = contention_ladder(vals, row["topN"])
        team_sums = sorted(stock_team_sums(pool, row["group"], row["stat"], row["topN"]))
        cap_j = percentile(team_sums, LUXURY_CAP_PERCENTILE)
        cap_ladder = percentile(sorted(ladder), LUXURY_CAP_PERCENTILE)
        derived.append(dict(row, teamSums=team_sums, ladder=ladder, capJuiced=cap_j,
                            capLadderAlt=cap_ladder))
        status = "ACTIVE" if row["enabled"] else "DISABLED"
        print(f"| {row['group']:<8} | {row['stat']:<4} | {row['topN']:>4} | {row['xblCap']:>7,.0f} | "
              f"{team_sums[0]:>9,.0f} | {percentile(team_sums, .5):>9,.1f} | {cap_j:>24,.1f} | "
              f"{team_sums[-1]:>9,.0f} | {cap_ladder:>12,.1f} | {cap_j / row['xblCap']:>6.3f} | {status} |")
    disabled = [d for d in derived if not d["enabled"]]
    if disabled:
        print("\nDISABLED rows remain:")
        for d in disabled:
            print(f"  - {d['group']}/{d['stat']}")
    else:
        print("\nDISABLED rows: none. DB1 closed the pitcher batterRatings gap, so the 8")
        print("rotation/bullpen POW-CON-SPD-FLD rows are active and derived from real stock-team")
        print("pitcher-batting distributions (F1 closed).")
    print("\nSensitivity (adopted stock-team caps at alternate percentiles, ACTIVE rows, for JK calibration):")
    print("| Group/Stat | 50th | 65th (default) | 75th | 90th |")
    print("|------------|------|----------------|------|------|")
    for d in derived:
        ts = d["teamSums"]
        print(f"| {d['group']}/{d['stat']:<4} | {percentile(ts, .5):>4,.0f} | {d['capJuiced']:>14,.1f} | "
              f"{percentile(ts, .75):>4,.0f} | {percentile(ts, .9):>4,.0f} |")
    return derived, sigma_j


def invert_mean_rating_ratio(pool, engine, scales):
    """First-order rating-scale per tier: invert the IV transform through the pool's
    attribute-cost composition. Empirical: find f such that scaling every rating by f
    scales total pool IV by s. Solved by bisection on the actual pool (deterministic)."""
    def scaled_total(f):
        tot = 0
        for p in pool:
            q = dict(p)
            q["bat"] = {k: v * f for k, v in p["bat"].items()}
            if p["isPitcher"]:
                q["pit"] = {k: v * f for k, v in p["pit"].items()}
                tot += engine.pool_iv(q)["attributes"]
            else:
                tot += engine.pool_iv(q)["attributes"]
        return tot

    # baseline at f=1 uses attribute IV only (traits/pitches excluded - they ride along)
    base_attr = scaled_total(1.0)
    out = {}
    for tier, s in scales.items():
        if s == 1.0:
            out[tier] = 1.0
            continue
        target = base_attr * s
        lo, hi = 0.4, 1.0
        for _ in range(40):
            mid = (lo + hi) / 2
            if scaled_total(mid) > target:
                hi = mid
            else:
                lo = mid
        out[tier] = (lo + hi) / 2
    return out


# ---------------------------------------------------------------------------
# R5 - EV-flatness across composed identities (§5.3 acceptance criterion)
# ---------------------------------------------------------------------------

BANDS = ["Power", "Contact", "Speed", "Defense", "Rotation", "Bullpen"]
BAND_STATS = {
    "Power": ["POW"], "Contact": ["CON"], "Speed": ["SPD"], "Defense": ["FLD", "ARM"],
    "Rotation": ["RVEL", "RJNK", "RACC"], "Bullpen": ["PVEL", "PJNK", "PACC"],
}
MOD_STAT_TO_LUX = {  # mod delta column -> (group, stat) of the luxury row it shifts
    "POW": ("hitters", "POW"), "CON": ("hitters", "CON"), "SPD": ("hitters", "SPD"),
    "FLD": ("hitters", "FLD"), "ARM": ("hitters", "ARM"),
    "RVEL": ("rotation", "VEL"), "RJNK": ("rotation", "JNK"), "RACC": ("rotation", "ACC"),
    "PVEL": ("bullpen", "VEL"), "PJNK": ("bullpen", "JNK"), "PACC": ("bullpen", "ACC"),
}


def band_scores(mods, mod_stats, xbl_caps):
    """Band-tag scores per §6.3, cap-normalized so '+130 FLD' is comparable to '+25 CON'.
    pos = positive deltas only (the band 'tag'); net = signed sum (harm included)."""
    out = {}
    for name, deltas in mods.items():
        pos, net = {}, {}
        for band in BANDS:
            pos[band] = sum(max(deltas[st], 0) / xbl_caps[st] for st in BAND_STATS[band])
            net[band] = sum(deltas[st] / xbl_caps[st] for st in BAND_STATS[band])
        out[name] = {"pos": pos, "net": net}
    return out


def compose_identity(priorities, scores, mods):
    """§6.3 greedy instantiation (T3; flagged for JK ahead of T8):

    INCREASES (2): score = sum_b priority_b x posScore_b  +  sum_b min(netScore_b, 0).
    Positive contribution counts only in priority bands; NEGATIVE side-deltas count
    everywhere (a legal roster necessarily carries every stat, so cap harm is real
    regardless of identity - this is what makes 'Call Your Shot' (-20 everything)
    unattractive to a rational Power team despite its +50 POW tag).
    Ties: total |delta| magnitude, then name (deterministic).

    DECREASES: '--' / '--'. The workbook makes decreases optional ('--' rows are legal;
    its own example team uses one), and under the sheet's mechanics a decrease is pure
    downside, so the rational stack skips them. Whether KBL should REQUIRE balanced
    inc/dec stacks is an open design point for JK (doc'd as amendment candidate)."""
    def magnitude(name):
        return sum(abs(v) for v in mods[name].values())

    def pick_increase(weight, taken):
        best_key, best_name = None, None
        for name, sc in scores.items():
            if name == "--" or name in taken:
                continue
            val = sum(weight[b] * sc["pos"][b] for b in BANDS)
            val += sum(min(sc["net"][b], 0) for b in BANDS)
            cand = (val, magnitude(name))
            if best_key is None or cand > best_key or (cand == best_key and name < best_name):
                best_key, best_name = cand, name
        return best_name

    # round-robin over priority bands (desc priority, name tiebreak) so a cross identity
    # gets one increase per band instead of both landing in whichever band's mods are
    # the largest cap-fractions (pitching mods dwarf hitting mods in the XBL tables)
    pri_bands = [b for b in sorted(BANDS, key=lambda b: (-priorities[b], b)) if priorities[b] > 0]
    if len(pri_bands) == 1:
        pri_bands = pri_bands * 2
    incs = []
    for b in pri_bands[:2]:
        w = {x: (priorities[x] if x == b else 0.0) for x in BANDS}
        incs.append(pick_increase(w, set(incs)))
    return {"increase": incs, "decrease": ["--", "--"]}


def identity_cap_shift(identity, mods, xbl_caps):
    """Net shift per mod stat as a FRACTION of the XBL cap (tier-invariant, §5.3)."""
    net = {st: 0.0 for st in MOD_STAT_TO_LUX}
    for name in identity["increase"]:
        for st in net:
            net[st] += mods[name][st]
    for name in identity["decrease"]:
        for st in net:
            net[st] -= mods[name][st]
    return {st: net[st] / xbl_caps[st] for st in net}


ROSTER_SHAPE = {
    # canonical 22-man stock-SMB4 shape: 13 hitters (8 fixed positions + 5 flex incl DH)
    # + 4 SP + 1 SP/RP + 3 RP + 1 CP  (matches the pool's per-team composition 440/20)
    "fixed_hitters": HITTER_POS,
    "flex_hitters": 5,
    "pitchers": [("SP", 4), ("SP/RP", 1), ("RP", 3), ("CP", 1)],
}


def roster_tax(roster, caps_by_rowkey, lux_rows):
    """Taxed-mode luxury bill for a 22-man roster (base ratings, §5.3)."""
    hitters = [p for p in roster if not p["isPitcher"]]
    rot = [p for p in roster if p["isPitcher"] and p["role"] in ("SP", "SP/RP")]
    pen = [p for p in roster if p["isPitcher"] and p["role"] in ("RP", "CP", "SP/RP")]  # JK ruling 2026-06-10: SP/RP counts toward pen concentration (T3-AUDIT MAJOR fix)
    total = 0.0
    binding = []
    for row in lux_rows:
        if not row.get("enabled", True):
            continue
        g, st, n = row["group"], row["stat"], row["topN"]
        if g == "hitters":
            vals = sorted((p["bat"][st] for p in hitters), reverse=True)[:n]
        elif g == "rotation":
            vals = sorted((p["pit"][st] if st in ("VEL", "JNK", "ACC") else p["bat"][st]
                           for p in rot), reverse=True)[:n]
        else:
            vals = sorted((p["pit"][st] if st in ("VEL", "JNK", "ACC") else p["bat"][st]
                           for p in pen), reverse=True)[:n]
        cap = max(caps_by_rowkey[(g, st)], 0.0)
        over = sum(vals) - cap
        if over > 0:
            tax = row["per100"] * (over / 100.0) ** row["curve"] + row["minAdder"]
            total += tax
            binding.append((f"{g}/{st}", over, tax))
    binding.sort(key=lambda b: -b[2])
    return total, binding


def greedy_build(pool, budget, caps_by_rowkey, lux_rows):
    """Best-achievable 22-man roster under budget incl. taxes (taxed balanceMode).
    Deterministic two-start hill climb (cheapest-feasible start + value-first start),
    best-improvement swaps; returns the higher-total-IV result."""
    hitters = sorted((p for p in pool if not p["isPitcher"]), key=lambda p: (-p["iv"], p["id"]))
    by_pos = {pos: [p for p in hitters if p["pos"] == pos] for pos in HITTER_POS}
    by_role = {role: sorted((p for p in pool if p["isPitcher"] and p["role"] == role),
                            key=lambda p: (-p["iv"], p["id"])) for role in PITCHER_ROLES}

    slots = [("pos", pos) for pos in ROSTER_SHAPE["fixed_hitters"]]
    slots += [("flex", None)] * ROSTER_SHAPE["flex_hitters"]
    for role, k in ROSTER_SHAPE["pitchers"]:
        slots += [("role", role)] * k

    def eligible(slot):
        kind, key = slot
        if kind == "pos":
            return by_pos[key]
        if kind == "flex":
            return hitters
        return by_role[key]

    def total_cost(roster):
        tax, _ = roster_tax(roster, caps_by_rowkey, lux_rows)
        return sum(p["iv"] for p in roster) + tax, tax

    def climb(assign):
        roster = list(assign)
        cost, _ = total_cost(roster)
        if cost > budget:  # repair: swap priciest slots down to the cheapest eligible
            order = sorted(range(len(slots)), key=lambda i: -roster[i]["iv"])
            for i in order:
                if cost <= budget:
                    break
                ids = {p["id"] for p in roster}
                cands = [c for c in eligible(slots[i]) if c["id"] not in ids]
                if not cands:
                    continue
                cheap = min(cands, key=lambda c: (c["iv"], c["id"]))
                trial = roster[:i] + [cheap] + roster[i + 1:]
                tcost, _ = total_cost(trial)
                if tcost < cost:  # accept any reduction; keep repairing
                    roster, cost = trial, tcost
        if cost > budget:
            return None
        improved = True
        while improved:
            improved = False
            best = None
            ids = {p["id"] for p in roster}
            for i, slot in enumerate(slots):
                for c in eligible(slot):
                    if c["id"] in ids or c["iv"] <= roster[i]["iv"]:
                        continue
                    trial = roster[:i] + [c] + roster[i + 1:]
                    tcost, _ = total_cost(trial)
                    if tcost <= budget:
                        gain = c["iv"] - roster[i]["iv"]
                        key = (gain, -tcost, c["id"])
                        if best is None or key > best[0]:
                            best = (key, i, c)
            if best is not None:
                _, i, c = best
                roster = roster[:i] + [c] + roster[i + 1:]
                improved = True
        return roster

    results = []
    # start A: cheapest feasible
    used = set()
    a = []
    for slot in slots:
        c = next(p for p in sorted(eligible(slot), key=lambda p: (p["iv"], p["id"])) if p["id"] not in used)
        a.append(c)
        used.add(c["id"])
    ra = climb(a)
    if ra:
        results.append(ra)
    # start B: value-first then repair
    used = set()
    b = []
    for slot in slots:
        c = next(p for p in eligible(slot) if p["id"] not in used)
        b.append(c)
        used.add(c["id"])
    rb = climb(b)
    if rb:
        results.append(rb)
    if not results:
        fail("greedy_build: no feasible roster under budget")
    return max(results, key=lambda r: sum(p["iv"] for p in r))


def r5_ev_flatness(pool, lux_derived, mods, mod_stats, tier_cap, sigma):
    print("\n" + "=" * 100)
    print("R5. EV-FLATNESS (§5.3 acceptance criterion) - best-achievable roster IV per composed")
    print(f"    identity, taxed mode, budget = juiced tierCap = {tier_cap:,.0f}, tolerance +/-{EV_FLATNESS_TOLERANCE:.0%}")
    print("=" * 100)
    xbl_caps = {st: next(d["xblCap"] for d in lux_derived
                         if (d["group"], d["stat"]) == MOD_STAT_TO_LUX[st]) for st in MOD_STAT_TO_LUX}
    scores = band_scores(mods, mod_stats, xbl_caps)
    neutral_caps = {(d["group"], d["stat"]): d["capJuiced"] for d in lux_derived}
    lux_scaled = [dict(r, per100=r["per100"] * sigma, minAdder=r["minAdder"] * sigma)
                  for r in lux_derived]

    identities = [(b, {x: (1.0 if x == b else 0.0) for x in BANDS}) for b in BANDS]
    crosses = [("Power+Rotation", ["Power", "Rotation"]), ("Contact+Speed", ["Contact", "Speed"]),
               ("Speed+Defense", ["Speed", "Defense"]), ("Defense+Rotation", ["Defense", "Rotation"]),
               ("Power+Bullpen", ["Power", "Bullpen"]), ("Contact+Defense", ["Contact", "Defense"])]
    for name, bands in crosses:
        identities.append((name, {x: (1.0 if x in bands else 0.0) for x in BANDS}))

    def run_identities(budget, label):
        rows = []
        for name, prio in identities:
            ident = compose_identity(prio, scores, mods)
            shift = identity_cap_shift(ident, mods, xbl_caps)
            caps = dict(neutral_caps)
            for st, frac in shift.items():
                key = MOD_STAT_TO_LUX[st]
                caps[key] = neutral_caps[key] * (1 + frac)
            roster = greedy_build(pool, budget, caps, lux_scaled)
            tot = sum(p["iv"] for p in roster)
            tax, binding = roster_tax(roster, caps, lux_scaled)
            rows.append({"name": name, "identity": ident, "iv": tot, "tax": tax,
                         "binding": binding[:3]})
        mean_iv = sum(r["iv"] for r in rows) / len(rows)
        print(f"\n### EV-flatness table - {label}\n")
        print("| Identity         | inc1 / inc2                       | dec1 / dec2 | roster IV   | taxes paid | IV vs mean | verdict |")
        print("|------------------|-----------------------------------|-------------|-------------|------------|------------|---------|")
        all_pass = True
        for r in rows:
            dev = r["iv"] / mean_iv - 1
            ok = abs(dev) <= EV_FLATNESS_TOLERANCE
            all_pass &= ok
            inc = " / ".join(r["identity"]["increase"])
            dec = " / ".join(r["identity"]["decrease"])
            print(f"| {r['name']:<16} | {inc:<33} | {dec:<11} | {r['iv']:>11,.0f} | {r['tax']:>10,.0f} | "
                  f"{dev:>+9.2%} | {'PASS' if ok else 'FAIL'}    |")
        print(f"\nCross-identity mean roster IV: {mean_iv:,.0f}  ->  "
              f"{'all identities within +/-10%' if all_pass else 'IDENTITIES OUTSIDE BAND - see above'}")
        for r in rows:
            if r["binding"]:
                b = ", ".join(f"{k} over {o:.0f} (${t:,.0f})" for k, o, t in r["binding"])
                print(f"  binding taxes - {r['name']}: {b}")
        return rows, mean_iv, all_pass

    rows, mean_iv, all_pass = run_identities(tier_cap, f"PRIMARY: budget = juiced tierCap = {tier_cap:,.0f}")
    print("""
NOTE (structural, doc'd in full): at construction time salary == IV, so 'maximize roster IV
under (payroll + taxes <= budget)' is exhausted by ANY full-budget tax-free roster - total
IV equals budget wherever the pool is deep enough to dodge every cap. Flatness at tierCap is
therefore structural; the table above is the §5.3 acceptance artifact, and the sensitivity
runs below show where the tax layer actually starts to shape outcomes.""")
    for mult in (1.5, 2.0):
        run_identities(tier_cap * mult, f"SENSITIVITY: budget = {mult:.1f} x tierCap = {tier_cap * mult:,.0f}")
    return rows, mean_iv, all_pass


# ---------------------------------------------------------------------------
# R6 - named-player spot checks
# ---------------------------------------------------------------------------

SPOT_CHECK_IDS = ["sir-longballo", "hrb-filthwick", "crc-fenomeno", "bee-pastimm", "wpg-drake", "blf-bradwick", "wdl-deals"]


def r6_spot_checks(pool, engine):
    print("\n" + "=" * 100)
    print("R6. SANITY NARRATIVE - named-player spot checks (full component breakdowns)")
    print("=" * 100 + "\n")
    by_id = {p["id"]: p for p in pool}
    ivs = sorted(p["iv"] for p in pool)
    ids = list(SPOT_CHECK_IDS)
    # always include the pool max and min for the narrative
    pool_sorted = sorted(pool, key=lambda p: (-p["iv"], p["id"]))
    for extreme in (pool_sorted[0]["id"], pool_sorted[-1]["id"]):
        if extreme not in ids:
            ids.append(extreme)
    for pid in ids:
        p = by_id.get(pid)
        if p is None:
            print(f"  (spot-check id {pid!r} not found - skipped)")
            continue
        parts = p.get("parts") or engine.pool_iv(p)
        pctl = sum(1 for v in ivs if v <= p["iv"]) / len(ivs)
        pos = p["role"] if p["isPitcher"] else p["pos"]
        raw_note = f", rawIV ${p['rawIV']:,}" if "rawIV" in p and p["rawIV"] != p["iv"] else ""
        print(f"- {p['name']} ({p['id']}, {pos}, grade {p['grade']}, bats {p['bats']}): kblIV ${p['iv']:,}{raw_note} (p{pctl:.0%})")
        if p["isPitcher"]:
            print(f"    components: pitch-attrs ${parts['pitchingAttributes']:,} | usage-bat/field ${parts['battingAttributes']:,} "
                  f"(bat {parts['hitterCurveBlock']} curves; FLD {parts['fieldingCurveBlock']} curve; w={parts['usageWeights']}) | traits ${parts['traits']:,} "
                  f"({', '.join(p['traits']) or 'none'}; two-way unlock ${parts['twoWayUnlock']:,}) | "
                  f"pitches ${parts['pitches']:,} | switch ${parts['handed']:,}")
        else:
            print(f"    components: attrs ${parts['attributes']:,} | traits ${parts['traits']:,} "
                  f"({', '.join(p['traits']) or 'none'}) | pitches ${parts['pitches']:,} | "
                  f"switch ${parts['handed']:,} | 2nd-pos ${parts['secondary']:,} ({p['secondary'] or '-'})")
        if p["isPitcher"]:
            print(f"    ratings: POW {p['bat']['POW']} CON {p['bat']['CON']} SPD {p['bat']['SPD']} "
                  f"FLD {p['bat']['FLD']} | VEL {p['pit']['VEL']} JNK {p['pit']['JNK']} ACC {p['pit']['ACC']} "
                  f"| arsenal {','.join(p['arsenal'])}")
        else:
            print(f"    ratings: POW {p['bat']['POW']} CON {p['bat']['CON']} SPD {p['bat']['SPD']} "
                  f"FLD {p['bat']['FLD']} ARM {p['bat']['ARM']}")
    by_id = {p["id"]: p for p in pool}
    oracle = [by_id["crc-fenomeno"], by_id["bee-pastimm"], by_id["wpg-drake"]]
    lad = by_id["blf-bradwick"]
    print("\nV117-FIX acceptance checks and reports (kblIV):")
    for p in oracle:
        print(f"  {p['name']}: ${p['iv']:,} (rawIV ${p['rawIV']:,})")
    lad_limit = 0.50 * lad["rawIV"]
    print(f"  {lad['name']}: ${lad['iv']:,} (rawIV ${lad['rawIV']:,})")
    if lad["iv"] > lad_limit:
        fail(f"V118 required crash failed: Lad Bradwick ${lad['iv']:,} > 50% of rawIV ${round(lad_limit):,}")
    print(f"  Lad crash gate: ${lad['iv']:,} <= ${round(lad_limit):,} PASS")
    print("\n  Fenomeno/Pastimm component bridge (report-only; parity band retired):")
    print("  | Component | Fenomeno | Pastimm | Pastimm - Fenomeno |")
    print("  |---|---:|---:|---:|")
    bridge_rows = [
        ("pitch attrs", "pitchingAttributes"),
        ("usage bat/field", "battingAttributes"),
        ("traits total", "traits"),
        ("two-way unlock", "twoWayUnlock"),
        ("pitches", "pitches"),
        ("arm slot", "angle"),
        ("total", "total"),
    ]
    fen_parts = oracle[0]["parts"]
    pas_parts = oracle[1]["parts"]
    for label, key in bridge_rows:
        f = fen_parts.get(key, 0)
        b = pas_parts.get(key, 0)
        print(f"  | {label} | ${f:,} | ${b:,} | ${b - f:+,} |")
    print(f"  Arm probe Pastimm vs Drake: ${oracle[1]['iv']:,} vs ${oracle[2]['iv']:,}; "
          f"trait-stack gap ${oracle[1]['iv'] - oracle[2]['iv']:,}")


# ---------------------------------------------------------------------------
# Emit src/data/tierParams.ts
# ---------------------------------------------------------------------------

def fmt(x, nd=4):
    if isinstance(x, float) and x.is_integer():
        return str(int(x))
    return f"{round(x, nd)}" if isinstance(x, float) else str(x)


def emit_tier_params(scales, farm, caps, lux_derived, sigma, mods, mod_stats, mean_iv, mean_ord, rating_scales):
    xbl_caps = {st: next(d["xblCap"] for d in lux_derived
                         if (d["group"], d["stat"]) == MOD_STAT_TO_LUX[st]) for st in MOD_STAT_TO_LUX}
    L = []
    a = L.append
    a("/**")
    a(" * tierParams.ts — T3 empirical tier parameters (DATA ONLY, generated).")
    a(" *")
    a(" * Generated by scripts/analyze-pool.py from:")
    a(" *   - src/data/playerDatabase.ts (440-player stock SMB4 pool = Juiced, spec D3)")
    a(" *   - src/data/ivCurves.ts + src/data/traitPricing.ts (T1, authoritative)")
    a(" *   - spec-docs/reference/Team_Builder_Archetype_Logic_Template.xlsx ('Luxury Cap' A:F + AT:BE)")
    a(" * DO NOT EDIT BY HAND - rerun the script. Deterministic: identical output every run.")
    a(" *")
    a(" * Derivations (full work shown in spec-docs/T3_POOL_ANALYSIS.md):")
    a(" *   - tier scales: grade-ladder method (§5.1) - median IV per SMB4 letter grade,")
    a(" *     piecewise-linear ladder; scale = ladder(meanOrdinal - steps) / ladder(meanOrdinal).")
    a(" *     Multiplicative transform: tierIV = juiced kblIV x scale.")
    a(" *   - V117 kblIV usage layer (§3.9): rawIV remains workbook-exact for anchors; pool")
    a(" *     analysis uses pitcher batting repriced on hitter curves with role usage weights.")
    a(" *   - tierCap = max(maxPoolIV / starBudgetShare, 22 x medianPoolIV x rosterHeadroom) (§5.2)")
    a(" *   - luxury caps: 65th percentile (luxuryCapPercentile) of the STOCK-TEAM top-N sum")
    a(" *     distribution - the 20 real SMB4 rosters' observed concentrations (§5.3; the")
    a(" *     ignore-budget contention ladder was derived too and REJECTED: caps so high the")
    a(" *     tax layer never binds - see T3_POOL_ANALYSIS.md §R4).")
    a(" *   - penalty $ scale sigma = pool median IV / XBL workbook anchor median salary;")
    a(" *     workbook curve shapes port unchanged except the approved quadratic pitcher")
    a(" *     rotation/bullpen POW/CON override (JK ruling 2026-07-15).")
    a(" *   - modification deltas: stored as FRACTIONS of the XBL cap they shift (tier-invariant),")
    a(" *     plus per-tier absolute tables (fraction x tier cap) (§5.3, §6.2).")
    a(" *   - farm nerf: one grade step left of the league tier (§7.4); no new free parameter.")
    a(" */")
    a("")
    a("export type TierKey = 'juiced' | 'standard' | 'nerfed';")
    a("")
    a("export interface TierShiftParams {")
    a("  scale: number;            // multiplicative IV-space transform vs the observed (Juiced) pool")
    a("  gradeStepsLeft: number;   // §5.1 definition of the tier")
    a("  impliedMeanIV: number;    // observed pool mean x scale")
    a("}")
    a("")
    a("export const TIER_SHIFTS: Record<TierKey, TierShiftParams> = {")
    for tier, steps in [("juiced", 0), ("standard", 1), ("nerfed", 2)]:
        a(f"  {tier}: {{ scale: {scales[tier]:.6f}, gradeStepsLeft: {steps}, "
          f"impliedMeanIV: {round(mean_iv * scales[tier])} }},")
    a("};")
    a("")
    a("/** First-order uniform rating multiplier that reproduces each tier's IV scale through")
    a(" *  the attribute curves on this pool (bisection-solved; Player Generator convenience). */")
    a("export const TIER_RATING_SCALES: Record<TierKey, number> = {")
    for tier in ["juiced", "standard", "nerfed"]:
        a(f"  {tier}: {rating_scales[tier]:.6f},")
    a("};")
    a("")
    a("/** §7.4 farm-draft nerf: farm pool IV scale RELATIVE TO its league tier (one grade step). */")
    a("export const FARM_NERF_SCALES: Record<TierKey, number> = {")
    for tier in ["juiced", "standard", "nerfed"]:
        a(f"  {tier}: {farm[tier]:.6f},")
    a("};")
    a("")
    a("export interface TierCapParams {")
    a("  tierCap: number;")
    a("  starBranch: number;       // maxPoolIV / starBudgetShare(0.33)")
    a("  rosterBranch: number;     // 22 x medianPoolIV x rosterHeadroom(1.15)")
    a("  maxPoolIV: number;")
    a("  medianPoolIV: number;")
    a("}")
    a("")
    a("export const TIER_CAPS: Record<TierKey, TierCapParams> = {")
    for tier in ["juiced", "standard", "nerfed"]:
        c = caps[tier]
        a(f"  {tier}: {{ tierCap: {round(c['cap'])}, starBranch: {round(c['starBranch'])}, "
          f"rosterBranch: {round(c['rosterBranch'])}, maxPoolIV: {round(c['maxIV'])}, "
          f"medianPoolIV: {round(c['medianIV'])} }},")
    a("};")
    a("")
    a("export interface LuxuryCapRow {")
    a("  group: 'hitters' | 'rotation' | 'bullpen';")
    a("  stat: 'POW' | 'CON' | 'SPD' | 'FLD' | 'ARM' | 'VEL' | 'JNK' | 'ACC';")
    a("  topN: number;")
    a("  cap: number;              // tier-scaled neutral cap (rating-sum)")
    a("  penaltyCurve: number;     // response shape; approved KBL tuning may supersede the XBL source curve")
    a("  penaltyPer100: number;    // $ per (overage/100)^curve - sigma-scaled to this pool")
    a("  minAdder: number;         // flat $ when over - sigma-scaled")
    a("}")
    a("")
    a("/**")
    a(" * Pitcher POW/CON are useful secondary skills, but should not price like dominant pitching.")
    a(" * Keep their stock-team-derived caps and dollar coefficients while using a soft quadratic")
    a(" * ramp so modest overages stay modest and only deliberate stacking becomes expensive.")
    a(" */")
    a(f"export const PITCHER_SECONDARY_BATTING_PENALTY_CURVE = {fmt(PITCHER_SECONDARY_BATTING_PENALTY_CURVE)};")
    a("")
    a("export const LUXURY_CAP_TABLES: Record<TierKey, LuxuryCapRow[]> = {")
    for tier in ["juiced", "standard", "nerfed"]:
        s = scales[tier]
        rs = rating_scales[tier]
        a(f"  {tier}: [")
        for d in lux_derived:
            if not d["enabled"]:
                continue
            cap = d["capJuiced"] * rs
            curve = ("PITCHER_SECONDARY_BATTING_PENALTY_CURVE"
                     if (d["group"], d["stat"]) in PITCHER_SECONDARY_BATTING_ROWS
                     else fmt(d["curve"]))
            a(f"    {{ group: '{d['group']}', stat: '{d['stat']}', topN: {d['topN']}, "
              f"cap: {round(cap, 1)}, penaltyCurve: {curve}, "
              f"penaltyPer100: {round(d['per100'] * sigma * s)}, minAdder: {round(d['minAdder'] * sigma * s)} }},")
        a("  ],")
    a("};")
    a("")
    a("/** Pitcher-BATTING luxury rows (rotation/bullpen POW CON SPD FLD) are ACTIVE after DB1.")
    a(" *  The stock pool now has 179/179 pitcher batterRatings, so no v1 luxury rows are disabled.")
    a(" *  Kept as an explicit empty registry for callers that check legacy disabled rows. */")
    a("export const DISABLED_LUXURY_ROWS: Array<Omit<LuxuryCapRow, 'cap'> & { xblCap: number; disabledReason: string }> = [")
    for d in lux_derived:
        if d["enabled"]:
            continue
        a(f"  {{ group: '{d['group']}', stat: '{d['stat']}', topN: {d['topN']}, "
          f"xblCap: {fmt(d['xblCap'])}, penaltyCurve: {fmt(d['curve'])}, "
          f"penaltyPer100: {round(d['per100'])}, minAdder: {round(d['minAdder'])}, "
          f"disabledReason: 'pitcher batterRatings data gap (89/178)' }},")
    a("];")
    a("")
    a("/** §6.2 modification deltas as FRACTIONS of the XBL cap of the luxury row each shifts.")
    a(" *  Tier-invariant ('+337 FLD' == +57.6% of the FLD cap at any tier). Apply as:")
    a(" *  shiftedCap = cap x (1 + sum(increases) - sum(decreases)). */")
    a("export type ModStat = 'POW' | 'CON' | 'SPD' | 'FLD' | 'ARM' | 'RVEL' | 'RJNK' | 'RACC' | 'PVEL' | 'PJNK' | 'PACC';")
    a("")
    a("export const CAP_MODIFICATION_FRACTIONS: Record<string, Record<ModStat, number>> = {")
    for name in mods:
        fr = ", ".join(f"{st}: {mods[name][st] / xbl_caps[st]:.6f}" for st in mod_stats)
        a(f"  {name!r}: {{ {fr} }},")
    a("};")
    a("")
    a("export const T3_DERIVATION_INPUTS = {")
    a(f"  poolSize: 440,")
    a(f"  poolMeanIV: {round(mean_iv)},")
    a(f"  poolMeanGradeOrdinal: {mean_ord:.4f},  // ladder D=0 .. S=11")
    a(f"  luxuryCapPercentile: {LUXURY_CAP_PERCENTILE},")
    a(f"  starBudgetShare: {STAR_BUDGET_SHARE},")
    a(f"  rosterHeadroom: {ROSTER_HEADROOM},")
    a(f"  penaltySigmaJuiced: {sigma:.6f},  // pool median IV / XBL anchor median salary")
    a(f"  contentionLadderTeams: {N_TEAMS},")
    a("} as const;")
    a("")
    with open(OUT_TS, "w") as f:
        f.write("\n".join(L))
    print(f"\n[write] src/data/tierParams.ts ({len(L)} lines)")


# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Analyze the KBL stock player pool and derive tier parameters.")
    parser.add_argument("--dump-oracle", help="Write a frozen T4 IV oracle JSON after the golden anchor gate passes.")
    args = parser.parse_args()

    curves = parse_curves()
    traits, pitches, switch, secpos, angles = parse_traits()
    pool, fa = parse_players()
    print(f"[load] pool: 440 stock players (20 teams x 22) + {len(fa)} free agents (excluded per spec D3/§5.1)")
    nobat = [p for p in pool if p["isPitcher"] and not p["hasBat"]]
    pitcher_count = sum(1 for p in pool if p["isPitcher"])
    sub_pitchers = [p for p in pool if p["isPitcher"] and p["armSlot"] == "Sub"]
    print(f"[load] F1 pitcher batterRatings coverage: {pitcher_count - len(nobat)}/{pitcher_count} present")
    print(f"[load] armSlot coverage: {pitcher_count}/{pitcher_count} present; Sub pitchers: "
          f"{len(sub_pitchers)} ({', '.join(p['name'] for p in sub_pitchers)})")
    if nobat:
        fail(f"F1 luxury-row flip requires complete pitcher batterRatings; missing {len(nobat)}")
    print(f"[load] curves: 18 blocks | traits: {len(traits)} | DB trait-name fixes applied: {TRAIT_NAME_FIXES}")

    engine = IVEngine(curves, traits, pitches, switch, secpos, angles)
    emit_usage_weights()

    # BOOTSTRAP GATE - abort before analysis if the workbook anchors don't reproduce
    anchors = load_anchors()
    anchor_count = run_anchor_gate(engine, anchors)
    anchors_median = percentile(sorted(a["expected"] for a in anchors), 0.5)

    for p in pool:
        p["rawParts"] = engine.pool_raw_iv(p)
        p["rawIV"] = p["rawParts"]["total"]
        p["parts"] = engine.pool_iv(p)
        p["iv"] = p["parts"]["total"]

    if args.dump_oracle:
        write_iv_oracle(args.dump_oracle, build_iv_oracle(engine, anchors, pool, anchor_count))
        return

    ivs = r1_distribution(pool)
    scales, farm, ladder, mean_ord, mean_iv = r2_tiers(pool, ivs)
    caps = r3_caps(ivs, scales)
    lux_rows, mods, mod_stats = load_luxury(WORKBOOK)
    lux_derived, sigma = r4_luxury(pool, lux_rows, mods, mod_stats, scales, anchors_median)
    rating_scales = invert_mean_rating_ratio(pool, engine, scales)
    print("\nFirst-order tier RATING multipliers (bisection on this pool's attribute IV):")
    for tier in ["juiced", "standard", "nerfed"]:
        print(f"  {tier}: x{rating_scales[tier]:.4f}")
    r5_ev_flatness(pool, lux_derived, mods, mod_stats, caps["juiced"]["cap"], sigma)
    r6_spot_checks(pool, engine)
    emit_tier_params(scales, farm, caps, lux_derived, sigma, mods, mod_stats, mean_iv, mean_ord, rating_scales)
    print("\nT3 analysis complete.")


if __name__ == "__main__":
    main()
