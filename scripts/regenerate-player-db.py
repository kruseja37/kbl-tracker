#!/usr/bin/env python3
"""Regenerate SMB4 team-rostered players from the roster source of truth.

This script intentionally rewrites only the team-rostered stock players inside
src/data/playerDatabase.ts. The free-agent block is preserved verbatim.
"""

from __future__ import annotations

import argparse
import ast
import csv
import hashlib
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
PLAYER_DB = ROOT / "src/data/playerDatabase.ts"
TRAIT_PRICING = ROOT / "src/data/traitPricing.ts"
WORKBOOK = ROOT / "reference-docs/SOURCE_OF_TRUTH_Super Mega Baseball 4 Rosters.xlsx"
PLAYERS_FINAL_CSV = ROOT / "spec-docs/data/players_final.csv"

TEAM_IDS_BY_SHEET = {
    "Beewolves": "beewolves",
    "Blowfish": "blowfish",
    "Buzzards": "buzzards",
    "Crocodons": "crocodons",
    "Freebooters": "freebooters",
    "Grapplers": "grapplers",
    "Heaters": "heaters",
    "Herbisaurs": "herbisaurs",
    "Hot Corners": "hot-corners",
    "Jacks": "jacks",
    "Moonstars": "moonstars",
    "Moose": "moose",
    "Nemesis": "nemesis",
    "Overdogs": "overdogs",
    "Platypi": "platypi",
    "Sandcats": "sand-cats",
    "Sawteeth": "sawteeth",
    "Sirloins": "sirloins",
    "Wideloads": "wideloads",
    "Wild Pigs": "wild-pigs",
}

SOT_NAME_TO_DB_NAME: dict[tuple[str, str], str] = {
    # Empty as of 2026-06-10: the four SOT name typos (Dano Yoshida, Seymour Scoks,
    # Lars Stadkeleef, Pex Flext) were corrected IN THE WORKBOOK per JK ruling
    # (DB1-AUDIT F3), so SOT names now match pre-regeneration DB names directly.
}

CHEMISTRY_NORMALIZATIONS = {
    "Competitve": "Competitive",
    "Spirted": "Spirited",
}

CHEMISTRY_CODE_FALLBACK = {
    "Competitive": "CMP",
    "Crafty": "CRA",
    "Disciplined": "DIS",
    "Scholarly": "SCH",
    "Spirited": "SPI",
}

TRAIT_NORMALIZATIONS = {
    "Base Rounds": "Base Rounder",
    "Clitch": "Clutch",
    "Con vs LHP": "CON vs LHP",
    "Con vs RHP": "CON vs RHP",
    "Con vs RPH": "CON vs RHP",
    "CON vs RPH": "CON vs RHP",
    "East Target": "Easy Target",
    "Elite 4": "Elite 4F",
    "K Neglecter": "K Neglector",
    "Off-speed Hitter": "Off-Speed Hitter",
    "POW vs PHP": "POW vs RHP",
    "PWR vs RHP": "POW vs RHP",
    "Slowpoke": "Slow Poke",
}

EMPTY_MARKERS = {"", "-", "None", "none", "NONE"}
ARM_SLOTS = {"High", "Mid", "Low", "Sub"}
OVERDOGS_CHEMISTRY_ANCHORS = {
    "Larry La'Joy": "Scholarly",
    "Chasey Kim": "Scholarly",
    "Werner Bergenberg": "Scholarly",
    "Carrie Wayward": "Crafty",
    "Slick Pickman": "Scholarly",
    "Brawn Thunderchump": "Scholarly",
    "David Diggler": "Spirited",
    "Rocket Ramon": "Spirited",
    "Doug Nerdwerd": "Spirited",
}
TRAIT_RULING_ANCHORS = {
    ("herbisaurs", "Gem Qualita"): ("Composed", None),
    ("overdogs", "Brawn Thunderchump"): ("Clutch", None),
    ("sand-cats", "Kara Kawaguchi"): ("Pinch Perfect", None),
}


@dataclass(frozen=True)
class CurrentPlayer:
    id: str
    name: str
    team_id: str
    gender: str


@dataclass(frozen=True)
class SotPlayer:
    team_sheet: str
    team_id: str
    name: str
    overall: str
    role: str
    is_pitcher: bool
    age: int
    bats: str
    throws: str
    chemistry_full: str
    trait1: str | None
    trait2: str | None
    primary_position: str
    secondary_position: str | None
    batter_ratings: dict[str, int]
    pitcher_role: str | None = None
    pitcher_ratings: dict[str, int] | None = None
    arsenal: tuple[str, ...] = ()
    arm_slot: str | None = None


class StopRegeneration(RuntimeError):
    pass


def ts_string(value: str) -> str:
    if "'" in value and '"' not in value:
        return '"' + value.replace("\\", "\\\\").replace('"', '\\"') + '"'
    return "'" + value.replace("\\", "\\\\").replace("'", "\\'") + "'"


def literal_from_ts(value: str) -> str:
    return ast.literal_eval(value)


def player_object_spans(players_body: str) -> Iterable[tuple[str, str]]:
    for player_id, obj, _raw_entry in player_entry_spans(players_body):
        yield player_id, obj


def player_entry_spans(players_body: str) -> Iterable[tuple[str, str, str]]:
    index = 0
    while True:
        match = re.search(r"\n\s*'([^']+)': \{", players_body[index:])
        if not match:
            return
        player_id = match.group(1)
        entry_start = index + match.start()
        object_start = index + match.end() - 1
        depth = 0
        cursor = object_start
        while cursor < len(players_body):
            char = players_body[cursor]
            if char == "{":
                depth += 1
            elif char == "}":
                depth -= 1
                if depth == 0:
                    break
            cursor += 1
        if depth != 0:
            raise StopRegeneration(f"Could not parse player object for {player_id}")
        entry_end = cursor + 1
        if entry_end < len(players_body) and players_body[entry_end] == ",":
            entry_end += 1
        raw_entry = players_body[entry_start + 1 : entry_end]
        yield player_id, players_body[object_start : cursor + 1], raw_entry
        index = entry_end


def parse_ts_string_prop(obj: str, prop: str) -> str | None:
    match = re.search(
        rf"\b{prop}:\s*('(?:\\'|[^'])*'|\"(?:\\\"|[^\"])*\")",
        obj,
    )
    return literal_from_ts(match.group(1)) if match else None


def parse_current_database(text: str) -> tuple[list[str], dict[str, CurrentPlayer], int]:
    teams_start = text.index("export const TEAMS")
    players_marker = "// ============================================\n// PLAYER DATA"
    teams_text = text[teams_start : text.index(players_marker)]
    team_order = [
        match.group(1)
        for match in re.finditer(r"\n\s*'([^']+)': \{\n\s*id: '[^']+'", teams_text)
        if match.group(1) != "free-agent"
    ]

    play_start = text.index("export const PLAYERS")
    opening = text.index("{", play_start)
    closing = text.index("\n};", opening)
    players_body = text[opening + 1 : closing]

    players: dict[str, CurrentPlayer] = {}
    free_agents = 0
    for player_id, obj in player_object_spans(players_body):
        name = parse_ts_string_prop(obj, "name")
        team_id = parse_ts_string_prop(obj, "teamId")
        gender = parse_ts_string_prop(obj, "gender")
        if not name or not team_id or not gender:
            raise StopRegeneration(f"Could not parse core fields for {player_id}")
        if team_id == "free-agent":
            free_agents += 1
            continue
        players.setdefault(player_id, CurrentPlayer(player_id, name, team_id, gender))
    return team_order, players, free_agents


def collect_free_agent_entries(text: str) -> list[str]:
    play_start = text.index("export const PLAYERS")
    opening = text.index("{", play_start)
    closing = text.index("\n};", opening)
    players_body = text[opening + 1 : closing]
    entries = []
    seen = set()
    for player_id, obj, raw_entry in player_entry_spans(players_body):
        if player_id in seen:
            continue
        seen.add(player_id)
        team_id = parse_ts_string_prop(obj, "teamId")
        if team_id == "free-agent":
            entries.append(raw_entry.rstrip())
    return entries


def normalize_chemistry(value: object, player_name: str) -> str:
    raw = "" if value is None else str(value).strip()
    raw = CHEMISTRY_NORMALIZATIONS.get(raw, raw)
    if raw in EMPTY_MARKERS:
        raise StopRegeneration(f"Missing chemistry for {player_name}")
    if raw not in CHEMISTRY_CODE_FALLBACK:
        raise StopRegeneration(f"Unrecognized chemistry {raw!r} for {player_name}")
    return raw


def normalize_trait(value: object, trait_names: set[str], player_name: str) -> str | None:
    raw = "" if value is None else str(value).strip()
    if raw in EMPTY_MARKERS:
        return None
    normalized = TRAIT_NORMALIZATIONS.get(raw, raw)
    if normalized not in trait_names:
        raise StopRegeneration(
            f"Unresolvable trait {raw!r} for {player_name}; normalized {normalized!r} is not in traitPricing.ts"
        )
    return normalized


def normalize_position(value: object) -> str | None:
    raw = "" if value is None else str(value).strip()
    return None if raw in EMPTY_MARKERS else raw


def int_cell(value: object, label: str, player_name: str) -> int:
    if value is None:
        raise StopRegeneration(f"Missing {label} for {player_name}")
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise StopRegeneration(f"Invalid {label} {value!r} for {player_name}") from exc


def load_trait_names() -> set[str]:
    text = TRAIT_PRICING.read_text()
    names = set(re.findall(r"name: '([^']+)'", text))
    if not names:
        raise StopRegeneration("No trait names parsed from traitPricing.ts")
    return names


def load_csv_chemistry_fallback() -> dict[tuple[str, str], str]:
    fallback: dict[tuple[str, str], str] = {}
    with PLAYERS_FINAL_CSV.open(newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            notes = (row.get("notes") or "").strip()
            if not notes.startswith("SMB4 "):
                continue
            team = notes.removeprefix("SMB4 ").strip()
            name = (row.get("name") or "").strip()
            chemistry = (row.get("chemistry") or "").strip()
            if team and name and chemistry:
                fallback[(team, name)] = CHEMISTRY_NORMALIZATIONS.get(chemistry, chemistry)
    return fallback


def find_pitcher_header(ws) -> int:
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 2).value).strip() == "Name" and str(ws.cell(row, 4).value).strip() == "Pow":
            return row
    raise StopRegeneration(f"Could not find pitcher header in sheet {ws.title}")


def load_sot_players(trait_names: set[str]) -> list[SotPlayer]:
    workbook = load_workbook(WORKBOOK, data_only=True)
    csv_chemistry = load_csv_chemistry_fallback()
    players: list[SotPlayer] = []

    for sheet_name, team_id in TEAM_IDS_BY_SHEET.items():
        if sheet_name not in workbook.sheetnames:
            raise StopRegeneration(f"Workbook missing sheet {sheet_name}")
        ws = workbook[sheet_name]
        pitcher_header = find_pitcher_header(ws)

        for row in range(2, pitcher_header):
            name_cell = ws.cell(row, 2).value
            overall_cell = ws.cell(row, 3).value
            if not name_cell or not overall_cell:
                continue
            name = str(name_cell).strip()
            if name.lower() == "name":
                continue
            role_marker = ws.cell(row, 1).value
            role = "BENCH" if str(role_marker).strip() == "BN" else "STARTER"
            primary = normalize_position(ws.cell(row, 4).value)
            if not primary:
                raise StopRegeneration(f"Missing primary position for {name}")
            chemistry_raw = ws.cell(row, 18).value or csv_chemistry.get((sheet_name, name))
            players.append(
                SotPlayer(
                    team_sheet=sheet_name,
                    team_id=team_id,
                    name=name,
                    overall=str(overall_cell).strip(),
                    role=role,
                    is_pitcher=False,
                    age=int_cell(ws.cell(row, 17).value, "age", name),
                    bats=str(ws.cell(row, 15).value).strip(),
                    throws=str(ws.cell(row, 16).value).strip(),
                    chemistry_full=normalize_chemistry(chemistry_raw, name),
                    trait1=normalize_trait(ws.cell(row, 13).value, trait_names, name),
                    trait2=normalize_trait(ws.cell(row, 14).value, trait_names, name),
                    primary_position=primary,
                    secondary_position=normalize_position(ws.cell(row, 5).value),
                    batter_ratings={
                        "power": int_cell(ws.cell(row, 6).value, "power", name),
                        "contact": int_cell(ws.cell(row, 7).value, "contact", name),
                        "speed": int_cell(ws.cell(row, 8).value, "speed", name),
                        "fielding": int_cell(ws.cell(row, 9).value, "fielding", name),
                        "arm": int_cell(ws.cell(row, 10).value, "arm", name),
                    },
                )
            )

        seen_pitcher = False
        bullpen = False
        for row in range(pitcher_header + 1, ws.max_row + 1):
            name_cell = ws.cell(row, 2).value
            overall_cell = ws.cell(row, 3).value
            if not name_cell or not overall_cell:
                if seen_pitcher:
                    bullpen = True
                continue
            name = str(name_cell).strip()
            if name.lower() == "name":
                continue
            seen_pitcher = True
            pitcher_role = str(ws.cell(row, 1).value).strip()
            if pitcher_role not in {"SP", "RP", "CP", "SP/RP"}:
                raise StopRegeneration(f"Invalid pitcher role {pitcher_role!r} for {name}")
            arm_slot = str(ws.cell(row, 12).value).strip()
            if arm_slot not in ARM_SLOTS:
                raise StopRegeneration(f"Invalid arm slot {arm_slot!r} for {name}")
            arsenal = tuple(part.strip() for part in str(ws.cell(row, 11).value).split(",") if part.strip())
            if not arsenal:
                raise StopRegeneration(f"Missing arsenal for pitcher {name}")
            chemistry_raw = ws.cell(row, 18).value or csv_chemistry.get((sheet_name, name))
            players.append(
                SotPlayer(
                    team_sheet=sheet_name,
                    team_id=team_id,
                    name=name,
                    overall=str(overall_cell).strip(),
                    role="BULLPEN" if bullpen else "ROTATION",
                    is_pitcher=True,
                    age=int_cell(ws.cell(row, 17).value, "age", name),
                    bats=str(ws.cell(row, 15).value).strip(),
                    throws=str(ws.cell(row, 16).value).strip(),
                    chemistry_full=normalize_chemistry(chemistry_raw, name),
                    trait1=normalize_trait(ws.cell(row, 13).value, trait_names, name),
                    trait2=normalize_trait(ws.cell(row, 14).value, trait_names, name),
                    primary_position="P",
                    secondary_position=None,
                    batter_ratings={
                        "power": int_cell(ws.cell(row, 4).value, "pitcher batting power", name),
                        "contact": int_cell(ws.cell(row, 5).value, "pitcher batting contact", name),
                        "speed": int_cell(ws.cell(row, 6).value, "pitcher batting speed", name),
                        "fielding": int_cell(ws.cell(row, 7).value, "pitcher fielding", name),
                        "arm": 0,
                    },
                    pitcher_role=pitcher_role,
                    pitcher_ratings={
                        "velocity": int_cell(ws.cell(row, 8).value, "velocity", name),
                        "junk": int_cell(ws.cell(row, 9).value, "junk", name),
                        "accuracy": int_cell(ws.cell(row, 10).value, "accuracy", name),
                    },
                    arsenal=arsenal,
                    arm_slot=arm_slot,
                )
            )
    return players


def build_id_mapping(current: dict[str, CurrentPlayer], sot_players: list[SotPlayer]):
    current_by_key: dict[tuple[str, str], CurrentPlayer] = {}
    for player in current.values():
        key = (player.team_id, player.name)
        if key in current_by_key:
            raise StopRegeneration(f"Duplicate current player key {key}")
        current_by_key[key] = player

    sot_to_current: dict[tuple[str, str], CurrentPlayer] = {}
    explicit_name_mappings: list[tuple[str, str, str]] = []

    for sot in sot_players:
        current_name = sot.name
        current_player = current_by_key.get((sot.team_id, current_name))
        configured_old_name = SOT_NAME_TO_DB_NAME.get((sot.team_id, sot.name))
        if not current_player and configured_old_name:
            current_name = SOT_NAME_TO_DB_NAME[(sot.team_id, sot.name)]
            current_player = current_by_key.get((sot.team_id, current_name))
        if not current_player:
            raise StopRegeneration(
                f"Unmapped SOT player {sot.team_sheet}/{sot.name}. Add an explicit mapping; no fuzzy matching is allowed."
            )
        sot_to_current[(sot.team_id, sot.name)] = current_player
        if configured_old_name:
            explicit_name_mappings.append((configured_old_name, sot.name, current_player.id))

    if len(sot_to_current) != len(sot_players):
        raise StopRegeneration("SOT-to-current mapping did not remain one-to-one")
    return sot_to_current, explicit_name_mappings


def derive_chemistry_mapping(
    current: dict[str, CurrentPlayer],
    sot_players: list[SotPlayer],
    sot_to_current: dict[tuple[str, str], CurrentPlayer],
    db_text: str,
) -> dict[str, str]:
    play_start = db_text.index("export const PLAYERS")
    opening = db_text.index("{", play_start)
    closing = db_text.index("\n};", opening)
    players_body = db_text[opening + 1 : closing]
    code_by_id: dict[str, str] = {}
    for player_id, obj in player_object_spans(players_body):
        team_id = parse_ts_string_prop(obj, "teamId")
        if team_id == "free-agent":
            continue
        code = parse_ts_string_prop(obj, "chemistry")
        if code:
            code_by_id[player_id] = code

    codes: dict[str, set[str]] = {}
    for sot in sot_players:
        current_player = sot_to_current[(sot.team_id, sot.name)]
        code = code_by_id.get(current_player.id)
        if code:
            codes.setdefault(sot.chemistry_full, set()).add(code)
    mapping = {}
    for full, observed in sorted(codes.items()):
        if len(observed) == 1:
            mapping[full] = next(iter(observed))
        else:
            fallback = CHEMISTRY_CODE_FALLBACK.get(full)
            if fallback not in observed:
                raise StopRegeneration(f"Ambiguous chemistry mapping for {full}: {sorted(observed)}")
            mapping[full] = fallback
    for full, code in CHEMISTRY_CODE_FALLBACK.items():
        mapping.setdefault(full, code)
    return mapping


def format_traits(trait1: str | None, trait2: str | None) -> str:
    if trait1 and trait2:
        return f"{{ trait1: {ts_string(trait1)}, trait2: {ts_string(trait2)} }}"
    if trait1:
        return f"{{ trait1: {ts_string(trait1)} }}"
    if trait2:
        return f"{{ trait1: {ts_string(trait2)} }}"
    return "{}"


def format_ratings(values: dict[str, int], keys: list[str]) -> str:
    return "{ " + ", ".join(f"{key}: {values[key]}" for key in keys) + " }"


def generate_player_object(player_id: str, current: CurrentPlayer, sot: SotPlayer, chemistry_code: str) -> str:
    lines = [
        f"  {ts_string(player_id)}: {{",
        f"    id: {ts_string(player_id)},",
        f"    name: {ts_string(sot.name)},",
        f"    teamId: {ts_string(sot.team_id)},",
        f"    age: {sot.age},",
        f"    gender: {ts_string(current.gender)},",
        f"    bats: {ts_string(sot.bats)},",
        f"    throws: {ts_string(sot.throws)},",
        f"    primaryPosition: {ts_string(sot.primary_position)},",
    ]
    if sot.secondary_position:
        lines.append(f"    secondaryPosition: {ts_string(sot.secondary_position)},")
    lines.extend(
        [
            f"    isPitcher: {'true' if sot.is_pitcher else 'false'},",
        ]
    )
    if sot.pitcher_role:
        lines.append(f"    pitcherRole: {ts_string(sot.pitcher_role)},")
    lines.extend(
        [
            f"    role: {ts_string(sot.role)},",
            f"    overall: {ts_string(sot.overall)},",
        ]
    )
    if sot.pitcher_ratings:
        lines.append(
            "    pitcherRatings: "
            + format_ratings(sot.pitcher_ratings, ["velocity", "junk", "accuracy"])
            + ","
        )
    lines.append(
        "    batterRatings: "
        + format_ratings(sot.batter_ratings, ["power", "contact", "speed", "fielding", "arm"])
        + ","
    )
    lines.extend(
        [
            f"    chemistry: {ts_string(chemistry_code)},",
            f"    traits: {format_traits(sot.trait1, sot.trait2)}",
        ]
    )
    if sot.is_pitcher:
        lines[-1] += ","
        lines.append("    arsenal: [" + ", ".join(ts_string(pitch) for pitch in sot.arsenal) + "],")
        lines.append(f"    armSlot: {ts_string(sot.arm_slot or '')}")
    lines.append("  }")
    return "\n".join(lines)


def role_heading(role: str) -> str:
    return {
        "STARTER": "POSITION PLAYERS (Starters)",
        "BENCH": "POSITION PLAYERS (Bench)",
        "ROTATION": "PITCHERS (Rotation)",
        "BULLPEN": "PITCHERS (Bullpen)",
    }[role]


def generate_team_block(
    team_order: list[str],
    team_names: dict[str, str],
    current: dict[str, CurrentPlayer],
    sot_players: list[SotPlayer],
    sot_to_current: dict[tuple[str, str], CurrentPlayer],
    chemistry_mapping: dict[str, str],
) -> str:
    players_by_team: dict[str, list[SotPlayer]] = {}
    for player in sot_players:
        players_by_team.setdefault(player.team_id, []).append(player)

    chunks: list[str] = []
    for team_id in team_order:
        players = players_by_team.get(team_id)
        if not players:
            continue
        display_name = team_names.get(team_id, team_id).upper()
        last_role = None
        for sot in players:
            if sot.role != last_role:
                if chunks:
                    chunks.append("")
                chunks.extend(
                    [
                        "  // ==========================================",
                        f"  // {display_name} - {role_heading(sot.role)}",
                        "  // ==========================================",
                    ]
                )
                last_role = sot.role
            current_player = sot_to_current[(sot.team_id, sot.name)]
            code = chemistry_mapping.get(sot.chemistry_full)
            if not code:
                raise StopRegeneration(f"No chemistry code for {sot.chemistry_full}")
            chunks.append(generate_player_object(current_player.id, current_player, sot, code) + ",")
    return "\n".join(chunks).rstrip() + "\n\n"


def parse_team_names_and_order(text: str) -> tuple[list[str], dict[str, str]]:
    teams_start = text.index("export const TEAMS")
    players_marker = "// ============================================\n// PLAYER DATA"
    teams_text = text[teams_start : text.index(players_marker)]
    team_order: list[str] = []
    team_names: dict[str, str] = {}
    for match in re.finditer(r"\n\s*'([^']+)': \{\n\s*id: '[^']+',\n\s*name: '([^']+)'", teams_text):
        team_id, name = match.groups()
        if team_id == "free-agent":
            continue
        team_order.append(team_id)
        team_names[team_id] = name
    return team_order, team_names


def ensure_arm_slot_interface(text: str) -> str:
    if "armSlot?: 'High' | 'Mid' | 'Low' | 'Sub';" in text:
        return text
    needle = "  arsenal?: string[];  // e.g., ['4F', '2F', 'CF', 'CB', 'SL', 'CH', 'FK']\n"
    replacement = needle + "  armSlot?: 'High' | 'Mid' | 'Low' | 'Sub';  // pitcher arm slot from SMB4 source roster\n"
    if needle not in text:
        raise StopRegeneration("Could not find arsenal interface line for armSlot insertion")
    return text.replace(needle, replacement, 1)


def generate_free_agent_block(free_agent_entries: list[str]) -> str:
    if not free_agent_entries:
        raise StopRegeneration("No free-agent entries found to preserve")
    return (
        "  // ==========================================\n"
        "  // FREE AGENTS\n"
        "  // ==========================================\n"
        + "\n".join(free_agent_entries)
        + "\n"
    )


def replace_players_body(text: str, new_body: str) -> str:
    play_start = text.index("export const PLAYERS")
    opening = text.index("{", play_start)
    closing = text.index("\n};", opening)
    return text[: opening + 2] + new_body + text[closing:]


def build_regenerated_text() -> tuple[str, dict[str, object]]:
    text = PLAYER_DB.read_text()
    trait_names = load_trait_names()
    team_order, current, _free_agent_count = parse_current_database(text)
    free_agent_entries = collect_free_agent_entries(text)
    parsed_order, team_names = parse_team_names_and_order(text)
    if parsed_order != team_order:
        raise StopRegeneration("Team order parser disagreement")
    sot_players = load_sot_players(trait_names)
    if len(sot_players) != 440:
        raise StopRegeneration(f"Expected 440 SOT team players, found {len(sot_players)}")
    if sum(1 for player in sot_players if player.is_pitcher) != 179:
        raise StopRegeneration("Expected 179 SOT pitchers")
    sot_to_current, explicit_name_mappings = build_id_mapping(current, sot_players)
    chemistry_mapping = derive_chemistry_mapping(current, sot_players, sot_to_current, text)
    team_block = generate_team_block(
        team_order,
        team_names,
        current,
        sot_players,
        sot_to_current,
        chemistry_mapping,
    )
    full_players_body = team_block + generate_free_agent_block(free_agent_entries)
    regenerated = ensure_arm_slot_interface(replace_players_body(text, full_players_body))
    return regenerated, {
        "chemistry_mapping": chemistry_mapping,
        "explicit_name_mappings": explicit_name_mappings,
        "free_agents": len(free_agent_entries),
        "sot_players": len(sot_players),
        "sot_pitchers": sum(1 for player in sot_players if player.is_pitcher),
    }


def parse_generated_players(text: str):
    play_start = text.index("export const PLAYERS")
    opening = text.index("{", play_start)
    closing = text.index("\n};", opening)
    players_body = text[opening + 1 : closing]
    parsed = {}
    for player_id, obj in player_object_spans(players_body):
        data: dict[str, object] = {
            "id": player_id,
            "name": parse_ts_string_prop(obj, "name"),
            "teamId": parse_ts_string_prop(obj, "teamId"),
            "age": int(re.search(r"\bage:\s*(\d+)", obj).group(1)),
            "bats": parse_ts_string_prop(obj, "bats"),
            "throws": parse_ts_string_prop(obj, "throws"),
            "primaryPosition": parse_ts_string_prop(obj, "primaryPosition"),
            "secondaryPosition": parse_ts_string_prop(obj, "secondaryPosition"),
            "role": parse_ts_string_prop(obj, "role"),
            "overall": parse_ts_string_prop(obj, "overall"),
            "chemistry": parse_ts_string_prop(obj, "chemistry"),
            "pitcherRole": parse_ts_string_prop(obj, "pitcherRole"),
            "armSlot": parse_ts_string_prop(obj, "armSlot"),
            "traits": {},
        }
        traits_match = re.search(r"\btraits:\s*\{([^}]*)\}", obj)
        if traits_match:
            for trait_key in ("trait1", "trait2"):
                value = parse_ts_string_prop("{" + traits_match.group(1) + "}", trait_key)
                if value:
                    data["traits"][trait_key] = value
        batter = re.search(
            r"batterRatings:\s*\{\s*power:\s*(\d+),\s*contact:\s*(\d+),\s*speed:\s*(\d+),\s*fielding:\s*(\d+),\s*arm:\s*(\d+)\s*\}",
            obj,
        )
        if batter:
            data["batterRatings"] = dict(zip(["power", "contact", "speed", "fielding", "arm"], map(int, batter.groups())))
        pitcher = re.search(
            r"pitcherRatings:\s*\{\s*velocity:\s*(\d+),\s*junk:\s*(\d+),\s*accuracy:\s*(\d+)\s*\}",
            obj,
        )
        if pitcher:
            data["pitcherRatings"] = dict(zip(["velocity", "junk", "accuracy"], map(int, pitcher.groups())))
        arsenal_match = re.search(r"arsenal:\s*\[([^\]]*)\]", obj)
        if arsenal_match:
            data["arsenal"] = re.findall(r"'([^']+)'", arsenal_match.group(1))
        parsed[player_id] = data
    return parsed


def verify_generated_database() -> list[str]:
    text = PLAYER_DB.read_text()
    trait_names = load_trait_names()
    _, current, _ = parse_current_database(text)
    sot_players = load_sot_players(trait_names)
    sot_to_current, _ = build_id_mapping(current, sot_players)
    chemistry_mapping = derive_chemistry_mapping(current, sot_players, sot_to_current, text)
    generated = parse_generated_players(text)
    team_players = [player for player in generated.values() if player["teamId"] != "free-agent"]
    pitchers = []
    rating_mismatches = []
    missing_batter = []
    missing_arsenal = []
    missing_arm_slot = []

    for sot in sot_players:
        current_player = sot_to_current[(sot.team_id, sot.name)]
        actual = generated[current_player.id]
        if actual["name"] != sot.name:
            rating_mismatches.append(f"{current_player.id} name {actual['name']} != {sot.name}")
        for field in ("age", "bats", "throws", "primaryPosition", "secondaryPosition", "role", "overall"):
            expected = {
                "age": sot.age,
                "bats": sot.bats,
                "throws": sot.throws,
                "primaryPosition": sot.primary_position,
                "secondaryPosition": sot.secondary_position,
                "role": sot.role,
                "overall": sot.overall,
            }[field]
            if actual.get(field) != expected:
                rating_mismatches.append(f"{current_player.id} {field} {actual.get(field)!r} != {expected!r}")
        if actual.get("chemistry") != chemistry_mapping[sot.chemistry_full]:
            rating_mismatches.append(f"{current_player.id} chemistry mismatch")
        for key, expected in sot.batter_ratings.items():
            actual_batter = actual.get("batterRatings") or {}
            if actual_batter.get(key) != expected:
                rating_mismatches.append(f"{current_player.id} batterRatings.{key} {actual_batter.get(key)} != {expected}")
        if sot.is_pitcher:
            pitchers.append(current_player.id)
            if not actual.get("batterRatings"):
                missing_batter.append(current_player.id)
            for key, expected in (sot.pitcher_ratings or {}).items():
                actual_pitcher = actual.get("pitcherRatings") or {}
                if actual_pitcher.get(key) != expected:
                    rating_mismatches.append(f"{current_player.id} pitcherRatings.{key} {actual_pitcher.get(key)} != {expected}")
            if actual.get("armSlot") != sot.arm_slot:
                missing_arm_slot.append(current_player.id)
            if not actual.get("arsenal"):
                missing_arsenal.append(current_player.id)

    if len(team_players) != 440:
        raise StopRegeneration(f"Verification failed: expected 440 team players, got {len(team_players)}")
    if missing_batter:
        raise StopRegeneration(f"Verification failed: missing pitcher batterRatings: {missing_batter}")
    if rating_mismatches:
        raise StopRegeneration("Verification failed: rating/field mismatches:\n" + "\n".join(rating_mismatches[:50]))
    if len(pitchers) != 179 or missing_arm_slot:
        raise StopRegeneration(f"Verification failed: armSlot coverage {179 - len(missing_arm_slot)}/179; missing {missing_arm_slot}")
    if missing_arsenal:
        raise StopRegeneration(f"Verification failed: missing pitcher arsenal: {missing_arsenal}")

    for name, full_chem in OVERDOGS_CHEMISTRY_ANCHORS.items():
        current_player = sot_to_current[("overdogs", name)]
        expected = chemistry_mapping[full_chem]
        actual = generated[current_player.id].get("chemistry")
        if actual != expected:
            raise StopRegeneration(f"Verification failed: Overdogs chemistry {name} {actual} != {expected}")

    for (team_id, name), expected_traits in TRAIT_RULING_ANCHORS.items():
        current_player = sot_to_current[(team_id, name)]
        traits = generated[current_player.id].get("traits") or {}
        actual_pair = (traits.get("trait1"), traits.get("trait2"))
        if actual_pair != expected_traits:
            raise StopRegeneration(f"Verification failed: trait ruling {name} {actual_pair} != {expected_traits}")

    return [
        "team players: 440",
        "pitchers with batterRatings: 179/179",
        "rating mismatches: 0",
        "pitcher armSlot coverage: 179/179",
        "pitcher arsenal coverage: 179/179",
        "Overdogs chemistry anchors: 9/9",
        "trait ruling anchors: 3/3",
    ]


def regenerate() -> int:
    original = PLAYER_DB.read_text()
    regenerated, report = build_regenerated_text()
    PLAYER_DB.write_text(regenerated)
    changed = hashlib.sha256(original.encode()).hexdigest() != hashlib.sha256(regenerated.encode()).hexdigest()
    print("Regenerated src/data/playerDatabase.ts")
    print(f"changed: {changed}")
    print(f"SOT team players: {report['sot_players']}")
    print(f"SOT pitchers: {report['sot_pitchers']}")
    print(f"free agents preserved: {report['free_agents']}")
    print("Chemistry mapping:")
    for full, code in sorted(report["chemistry_mapping"].items()):
        print(f"  {full} -> {code}")
    print("Explicit ID-preservation name mappings:")
    for old_name, sot_name, player_id in report["explicit_name_mappings"]:
        print(f"  {old_name} -> {sot_name} -> {player_id}")
    print("Trait normalizations:")
    for raw, normalized in sorted(TRAIT_NORMALIZATIONS.items()):
        print(f"  {raw} -> {normalized}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--verify", action="store_true", help="verify regenerated playerDatabase.ts against the SOT")
    args = parser.parse_args()
    try:
        if args.verify:
            for line in verify_generated_database():
                print(line)
            print("verification: PASS")
            return 0
        return regenerate()
    except StopRegeneration as exc:
        print(f"BLOCKED: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
